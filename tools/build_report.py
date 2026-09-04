# -*- coding: utf-8 -*-
"""Интерактивный отчёт для родителей. Собирается ИЗ журнала кассы, руками не правится."""
import datetime, json, re
from openpyxl import load_workbook

import journal

import os
SRC = os.environ.get("SRC", "/home/claude/Касса_2В_2026-2027.xlsx")
OUT = os.environ.get("OUT", "/home/claude/index.html")
PASSWORD = os.environ.get("PASSWORD", "")
ALLOW_UNSAFE = os.environ.get("ALLOW_UNSAFE") == "1"

# Предохранитель. Маскировки имён больше нет (см. pub()), и шифрование осталось
# единственным, что закрывает страницу от постороннего: с пустым PASSWORD в файл
# легли бы открытым текстом фамилия, имя, день рождения и долг каждого ребёнка.
# Раньше в паре с этой проверкой стояла вторая, на MASK; после возврата полных
# имён флаг MASK перестал что-либо обезличивать и убран - оставленный, он врал бы
# следующему читателю про вторую линию защиты, которой нет.
# Проверка идёт до чтения журнала. ALLOW_UNSAFE=1 снимает её: только для именной
# версии, которая остаётся в папке Свод на Google Диске и в репозиторий не идёт.
if not ALLOW_UNSAFE and not PASSWORD:
    raise SystemExit(
        "СБОРКА ОСТАНОВЛЕНА: PASSWORD пуст.\n"
        "Страница собралась бы в открытом виде — payload лежал бы в HTML как есть,\n"
        "и все данные прочитал бы любой, кто открыл ссылку или скачал файл:\n"
        "имена детей на странице полные, шифрование - единственное, что их закрывает.\n"
        "Задайте PASSWORD='кодовое-слово' либо, если версия намеренно не для\n"
        "публикации, соберите с ALLOW_UNSAFE=1.")

_as_of = os.environ.get("AS_OF", "").strip()
try:
    AS_OF = datetime.date.fromisoformat(_as_of) if _as_of else datetime.date.today()
except ValueError:
    raise SystemExit(f"AS_OF={_as_of!r}: ожидается дата в формате ГГГГ-ММ-ДД, например 2026-08-07.")

STAMP = os.environ.get("STAMP") or datetime.datetime.now().strftime("%d.%m.%Y %H:%M")


PBKDF2_ITERS = 600000   # подставляется и в страницу через __ITERS__: расходиться нельзя


def norm_pass(p: str) -> str:
    """Нормализация кода перед выводом ключа: регистр не важен, краевые пробелы срезаются.
    Родители набирают код на телефоне, где включена автозаглавная буква и легко
    прилетает пробел из буфера. Ровно то же делает decryptPayload на странице -
    иначе сборка и страница выведут разные ключи и отчёт не откроется."""
    return p.strip().lower()


def encrypt_payload(plaintext: str, password: str) -> str:
    """AES-256-GCM, ключ из PBKDF2-SHA256. Возвращает base64(salt|iv|ciphertext)."""
    import base64
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
    salt = os.urandom(16)               # новые при каждой сборке: повтор nonce ломает GCM
    iv = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=PBKDF2_ITERS)
    key = kdf.derive(norm_pass(password).encode())
    ct = AESGCM(key).encrypt(iv, plaintext.encode(), None)
    return base64.b64encode(salt + iv + ct).decode()


def pub(n):
    """Имя для страницы. Точка, где решается формат: раньше отсюда выходило
    «Артур Е.», теперь - колонка C журнала как есть, «Елисеев Артур».

    Решение родкома от 27.08.2026: с 26 детьми подпись «Артур Е.» неудобно искать
    глазами в списке. Риск обсуждали и приняли - код доступа знают все семьи
    класса, и при его утечке наружу уходят фамилия, имя, день рождения и долг
    каждого ребёнка. Формат «Фамилия Имя» повторяет журнал: список отсортирован
    по фамилии, и страница читается в том же порядке, что бумажный список.

    Функция оставлена именно как эта точка. Менять формат - здесь, в одном месте,
    а не в полудюжине мест вёрстки.
    """
    return str(n or "").strip()


# Подпись казначея на странице. Осталась короткой намеренно: решение родкома
# касалось списка детей, про казначея речи не было, а его фамилия - такие же
# персональные данные третьего лица. Одна строка, если решат иначе.
TREASURER_LABEL = "Анна Е."


wb = load_workbook(SRC, data_only=True)
sv, sb, uch, dgy, dgs, uc, vz, rs = (wb["Свод"], wb["Сборы"], wb["Ученики"], wb["Долги_год"],
                                     wb["Долги_срок"], wb["Участие"], wb["Взносы"], wb["Расходы"])
gr = wb["График"]
# Обещанное, но не оплаченное. Лист появился позже остальных, поэтому его
# отсутствие - не ошибка: у прошлогоднего журнала его просто нет.
ob = wb["Обязательства"] if "Обязательства" in wb.sheetnames else None
# Планируемые события класса. Деньги за них идут мимо кассы - родитель платит
# организатору напрямую, - поэтому лист не участвует ни в одном расчёте: ни в
# поступлениях, ни в расходах, ни в долях, ни в долгах. На страницу он выходит
# только календарём. Лист появился позже остальных, его отсутствие - не ошибка.
ev = wb["События"] if "События" in wb.sheetnames else None
# Личные списания: часть уже сделанного расхода, адресованная конкретному ребёнку.
# Сама трата целиком лежит в «Расходах» и входит в «потрачено» - этот лист её не
# добавляет, а только делит иначе: не поровну на класс, а на того, кому вещь
# досталась. Лист появился позже остальных, его отсутствие - не ошибка.
ls = wb["Личные списания"] if "Личные списания" in wb.sheetnames else None
# Хронология кассы со сквозным остатком: приход и расход по дням, сверху вниз.
# Родитель видит «собрано», «потрачено» и остаток, но свести их сам не может -
# взносы и траты лежат в разных списках. Этот лист даёт пройти сверху вниз и
# прийти к той же цифре. Лист появился позже остальных, его отсутствие - не ошибка.
mv = wb["Движение"] if "Движение" in wb.sheetnames else None

# Границы данных не записаны числами: их читает journal.bounds() из самого журнала -
# по шапке сверху и по строке «ИТОГО…» снизу. Класс растёт посреди года, под списки
# в журнале расширяют запас строк, и константа в этот день молча теряет хвост списка.
# Пустой диапазон для отсутствующих «Обязательств» - тот же случай, что лист без строк.
S_FIRST, S_LAST = journal.bounds(uch, "№")
B_FIRST, B_LAST = journal.bounds(sb, "Код")
GR_FIRST, GR_LAST = journal.bounds(gr, "Сбор")
VZ_FIRST, VZ_LAST = journal.bounds(vz, "Дата")
RS_FIRST, RS_LAST = journal.bounds(rs, "Дата")
OB_FIRST, OB_LAST = journal.bounds(ob, "Срок") if ob is not None else (0, -1)
# У «Событий» строки «ИТОГО» нет: суммы там с одного участника и в одну сумму не
# складываются. Нижняя граница - конец листа, см. journal.bounds_open().
EV_FIRST, EV_LAST = journal.bounds_open(ev, "Дата") if ev is not None else (0, -1)
# У «Личных списаний» строка «ИТОГО списано лично» есть - границы читаются обычным
# bounds(). Пустой лист даёт пустой диапазон, и на странице не появляется ничего.
LS_FIRST, LS_LAST = journal.bounds(ls, "Дата") if ls is not None else (0, -1)
# У «Движения» итога нет: приход с расходом в одну сумму не складываются. Данные
# кончаются первой пустой датой - ниже лежит блок сверки и подпись под ним,
# см. journal.bounds_blank().
MV_FIRST, MV_LAST = journal.bounds_blank(mv, "Дата") if mv is not None else (0, -1)
# Колонка ребёнка в матрицах выводится из номера его строки, поэтому matrix_col
# заодно сверяет имена в шапках со списком: разъехавшись, они выдали бы детям
# чужие долги, и отчёт всё равно собрался бы.
MC = journal.matrix_col(uch, S_FIRST, S_LAST, (uc, dgy, dgs))


def dt(x):
    """Дата из ячейки. Проверка именно на дату, а не на наличие strftime: после
    пересчёта формул LibreOffice пустой результат MINIFS приходит из ячейки как
    datetime.time(0, 0), у которого strftime тоже есть, и пустой срок превращался
    в «01.01.1900» и уезжал в календарь событий."""
    return x.strftime("%d.%m.%Y") if isinstance(x, (datetime.datetime, datetime.date)) else ""


def num(x):
    return x if isinstance(x, (int, float)) else 0


def bday(x):
    """День и месяц из колонки K. В журнале там текст «07.03.2018», но стоит один
    раз ввести дату не текстом, и Таблицы отдадут настоящую дату: str() даст
    «2026-08-30 00:00:00», срез - «2026-», и в календаре появится NaN. Поэтому
    дата разбирается как дата, а из текста берётся только «ДД.ММ» и только если
    это действительно «ДД.ММ»: пустой MINIFS приходит как time(0, 0), у которого
    срез дал бы «00:00»."""
    if isinstance(x, (datetime.datetime, datetime.date)):
        return x.strftime("%d.%m")
    s = str(x or "").strip()[:5]
    return s if re.fullmatch(r"\d{2}\.\d{2}", s) else ""


# Подпись строки учителя на странице. Имя из колонки C журнала наружу не идёт:
# там может оказаться настоящее ФИО, а публичная версия его не показывает.
# Это подпись элемента интерфейса, а не данные: дата рождения и сам факт строки
# лежат в журнале. tools/verify.py читает эту строку отсюда - она единственная,
# которой техническая строка вправе совпасть с содержимым страницы.
TEACHER_LABEL = "Учитель класса"


def schedule(title):
    """График платежей сбора с листа «График»: B сбор, C дата, D сумма, E комментарий.
    Порядок в журнале не гарантирован, поэтому сортируем по дате сами."""
    out = []
    for r in range(GR_FIRST, GR_LAST + 1):
        if gr.cell(row=r, column=2).value != title:
            continue
        d, amount = gr.cell(row=r, column=3).value, num(gr.cell(row=r, column=4).value)
        if not d or not amount:
            continue
        out.append({"date": dt(d), "amount": amount,
                    "note": gr.cell(row=r, column=5).value or ""})
    return sorted(out, key=lambda x: x["date"].split(".")[::-1])


def money(x):
    """«1000» -> «1 000» неразрывным пробелом, как это делает rub() на странице."""
    return f"{int(x):,}".replace(",", " ")


_MISSING = object()


def svod(label, default=_MISSING):
    """Итог с листа «Свод» по подписи в колонке B, а не по номеру строки.
    Строки на листе сдвигаются - так между «Остатком кассы» и «Ждём к сроку»
    появились «Обещано» и «Свободный остаток», - а подписи остаются прежними.
    Без default отсутствие строки останавливает сборку: молча подставленный
    ноль уехал бы в шапку и выглядел бы как настоящее число."""
    for r in range(1, sv.max_row + 1):
        if str(sv.cell(row=r, column=2).value or "").strip() == label:
            return num(sv.cell(row=r, column=3).value)
    if default is _MISSING:
        raise SystemExit(
            f"СБОРКА ОСТАНОВЛЕНА: на листе «Свод» нет строки «{label}».\n"
            "Итоги читаются по подписям в колонке B - проверьте, не переименована ли строка.")
    return default


totals = {"collected": svod("Поступило всего"), "spent": svod("Потрачено всего"),
          "rest": svod("Остаток кассы"), "dueNow": svod("Ждём к ближайшему сроку"),
          "dueYear": svod("Ждём за год всего"),
          # Обязательств в старом журнале нет: тогда обещано ноль,
          # а свободный остаток равен остатку кассы.
          "promised": svod("Обещано, но не оплачено", 0)}
totals["free"] = svod("Свободный остаток", totals["rest"] - totals["promised"])

# Обещано, но ещё не оплачено: B срок, C сбор, D на что, E сумма, F кому, G комментарий.
# Пустая строка опознаётся по пустой сумме - как на листах «Взносы» и «Расходы».
promises = []
if ob is not None:
    for r in range(OB_FIRST, OB_LAST + 1):
        if not ob.cell(row=r, column=5).value:
            continue
        promises.append({"date": dt(ob.cell(row=r, column=2).value),
                         "sbor": ob.cell(row=r, column=3).value or "",
                         "what": ob.cell(row=r, column=4).value or "",
                         "amount": num(ob.cell(row=r, column=5).value),
                         "who": ob.cell(row=r, column=6).value or "",
                         "note": ob.cell(row=r, column=7).value or ""})
    promises.sort(key=lambda x: x["date"].split(".")[::-1])


# Движение денег: B дата, C операция, D пришло, E ушло, F остаток после операции,
# G подтверждение. Лист считается из «Взносов» и «Расходов», руками не правится;
# взносы одного дня уже сведены там в одну строку («Взносы, 10 чел.»).
#
# Возмещение от соседнего класса стоит здесь в приходе, хотя в «Расходах» это
# строка со знаком минус. Это не противоречие: в «Расходах» знак нужен, чтобы
# возврат уменьшил и общие траты, и долю каждого ребёнка, а в ленте кассы деньги
# именно пришли - в этот день на карте стало больше. Так и задумано.
flow = []
for r in range(MV_FIRST, MV_LAST + 1):
    date = dt(mv.cell(row=r, column=2).value)
    if not date:
        continue
    flow.append({"date": date, "what": mv.cell(row=r, column=3).value or "",
                 "in": num(mv.cell(row=r, column=4).value),
                 "out": num(mv.cell(row=r, column=5).value),
                 "bal": num(mv.cell(row=r, column=6).value),
                 "proof": str(mv.cell(row=r, column=7).value or "со слов").lower()})

# Вся ценность ленты в том, что по ней можно пройти сверху вниз и прийти к остатку
# из шапки. Значит проверяются обе вещи: каждый шаг (остаток предыдущей строки плюс
# приход минус расход) и последняя строка против «Остатка кассы» со «Свода». Не
# сошлось - сборка останавливается: страница, по которой нельзя сойтись, хуже, чем
# её отсутствие, потому что выглядит как доказательство.
_prev = 0
for i, f in enumerate(flow):
    want = _prev + f["in"] - f["out"]
    if abs(want - f["bal"]) > 0.005:
        raise SystemExit(
            f"СБОРКА ОСТАНОВЛЕНА: на листе «Движение» в строке {MV_FIRST + i} остаток не сходится.\n"
            f"«{f['what']}» от {f['date']}: {_prev:.2f} + {f['in']:.2f} - {f['out']:.2f} = {want:.2f}, "
            f"а в колонке F стоит {f['bal']:.2f}.")
    _prev = f["bal"]
if flow and abs(_prev - totals["rest"]) > 0.005:
    raise SystemExit(
        f"СБОРКА ОСТАНОВЛЕНА: последний остаток «Движения» ({_prev:.2f}) не сходится "
        f"с остатком кассы со «Свода» ({totals['rest']:.2f}).\n"
        "Лента для того и нужна, чтобы родитель пришёл по ней к той же цифре, "
        "что в шапке - пересчитайте журнал.")


# Планируемые события: B дата, C что, D сумма с участника, E кто организует,
# F комментарий. Заполненной считается строка с датой и названием: сумма бывает
# пустой (событие объявлено, цену ещё не назвали), и такую строку в календаре
# показать всё равно надо - без суммы в заголовке.
#
# Цена держится отдельным полем, а не приклеена к названию: в клетке месяца она
# идёт одной строкой с ним, а в «Ближайшем месяце» уходит вниз, к пояснению.
# Склеенный заголовок «Фотографирование класса - 650 ₽ с человека» разгонял
# карточку на телефоне в четыре строки, ломая ритм списка.
#
# Пояснение от кода («оплата напрямую…») дописывается ТОЛЬКО если про кассу
# ничего не сказано в комментарии журнала. Казначей обычно пишет это сам, и тогда
# две одинаковые фразы подряд читаются как машинная приписка - доверия к странице
# они не добавляют. Совсем убрать пояснение нельзя: метка «платят сами» стоит
# рядом с суммой, а сумма в календаре читается как «сдать в кассу», и этот вопрос
# нельзя оставлять на то, заполнила ли казначей нужную ячейку.
plans = []
for r in range(EV_FIRST, EV_LAST + 1):
    date, what = dt(ev.cell(row=r, column=2).value), ev.cell(row=r, column=3).value
    if not date or not what:
        continue
    amount = ev.cell(row=r, column=4).value
    who = str(ev.cell(row=r, column=5).value or "").strip()
    comment = str(ev.cell(row=r, column=6).value or "").strip()
    parts = [f"Организует: {who}" if who else "", comment.rstrip(". ")]
    if "касс" not in comment.lower():
        parts.append("Оплата напрямую организатору, в кассу класса эти деньги не идут")
    plans.append({"date": date, "kind": "plan", "title": str(what).strip(),
                  "amt": (f"{money(amount)} ₽ с человека"
                          if isinstance(amount, (int, float)) and amount else ""),
                  "note": " · ".join([p for p in parts if p])})
plans.sort(key=lambda x: x["date"].split(".")[::-1])


kids = []
for r in range(S_FIRST, S_LAST + 1):
    n = uch.cell(row=r, column=3).value
    if not n:
        continue
    # Колонка L - комментарий. «Не ученик…» помечает строку, которая ребёнком класса
    # не считается - ни в чипах, ни в знаменателе доли расходов, ни в основной таблице.
    # Таких строк два вида, и различает их ключевое слово в той же пометке:
    # внешний плательщик (соседний класс, делящий кабинет) - его деньги лежат в кассе
    # и показаны отдельным блоком; и учитель - у него нет ни взносов, ни остатка,
    # строка нужна только ради дня рождения в календаре.
    note = str(uch.cell(row=r, column=12).value or "").strip().lower()
    ext = note.startswith("не ученик")
    teach = ext and "учитель" in note
    # ord - позиция в журнале, то есть порядок по фамилии. Наружу уходит только число:
    # оно служит скрытым ключом сортировки и вторичным ключом при равных суммах.
    kids.append({"row": r, "col": MC + (r - S_FIRST), "ord": len(kids) + 1, "raw": n,
                 # У учителя на странице подпись из кода, а не имя из журнала:
                 # в колонке C может оказаться настоящее ФИО.
                 "name": TEACHER_LABEL if teach else pub(n),
                 "ext": ext, "teach": teach,
                 "prev": num(uch.cell(row=r, column=4).value),
                 "paid": num(uch.cell(row=r, column=5).value),
                 "debtY": num(uch.cell(row=r, column=7).value),
                 "debtN": num(uch.cell(row=r, column=8).value),
                 "share": num(uch.cell(row=r, column=9).value),
                 "rest": num(uch.cell(row=r, column=10).value),
                 "bday": bday(uch.cell(row=r, column=11).value),
                 # Личные списания ребёнка - заполняются ниже, с листа
                 # «Личные списания»; у большинства детей список пустой.
                 "own": [], "ownSum": 0,
                 "by": []})

# Личные списания: B дата, C сбор, D ребёнок, E за что, F сумма, G подтверждение,
# H комментарий. Заполненной считается строка, где есть и ребёнок, и сумма -
# остальное бывает пустым, а без этих двух полей адресовать нечего и некому.
#
# Имя сверяется со списком «Учеников» и незнакомое останавливает сборку. Тихо
# пропущенная строка - это ровно тот случай, ради которого лист и заведён: доля
# в журнале уже посчитана с личным списанием, а на странице его бы не было, и
# родитель увидел бы долю, которая ниоткуда не складывается.
_by_name = {k["raw"]: k for k in kids}
for r in range(LS_FIRST, LS_LAST + 1):
    who = str(ls.cell(row=r, column=4).value or "").strip()
    amount = num(ls.cell(row=r, column=6).value)
    if not who or not amount:
        continue
    k = _by_name.get(who)
    if k is None:
        raise SystemExit(
            f"СБОРКА ОСТАНОВЛЕНА: на листе «Личные списания» в строке {r} стоит «{who}», "
            "а в списке «Ученики» такого имени нет.\n"
            "Имя ребёнка должно совпадать со списком слово в слово - иначе списание "
            "не к кому отнести, а доля расходов в журнале его уже учитывает.")
    k["own"].append({"date": dt(ls.cell(row=r, column=2).value),
                     "sbor": str(ls.cell(row=r, column=3).value or "").strip(),
                     "what": ls.cell(row=r, column=5).value or "",
                     "amount": amount,
                     "proof": str(ls.cell(row=r, column=7).value or "со слов").lower(),
                     "note": ls.cell(row=r, column=8).value or ""})
for k in kids:
    k["own"].sort(key=lambda x: x["date"].split(".")[::-1])
    k["ownSum"] = sum(o["amount"] for o in k["own"])

sbory = []
for r in range(B_FIRST, B_LAST + 1):
    title = sb.cell(row=r, column=3).value
    if not title:
        continue
    per, first = num(sb.cell(row=r, column=4).value), num(sb.cell(row=r, column=5).value)
    # Колонка O - расход на участника: общая часть, делённая поровну. Личные
    # списания в неё не входят, они добавляются каждому ребёнку своим числом,
    # иначе доля в таблице сбора разошлась бы с долей из «Учеников».
    share = num(sb.cell(row=r, column=15).value)
    parts = []
    for k in kids:
        if uc.cell(row=r, column=k["col"]).value != 1:
            continue
        p = sum(num(vz.cell(row=vr, column=5).value) for vr in range(VZ_FIRST, VZ_LAST + 1)
                if vz.cell(row=vr, column=3).value == k["raw"]
                and vz.cell(row=vr, column=4).value == title)
        dy, dn = num(dgy.cell(row=r, column=k["col"]).value), num(dgs.cell(row=r, column=k["col"]).value)
        own = sum(o["amount"] for o in k["own"] if o["sbor"] == title)
        sh = share + own
        parts.append({"ord": k["ord"], "name": k["name"], "paid": p,
                      "debtY": dy, "debtN": dn, "rest": p - sh})
        k["by"].append({"sbor": title, "code": sb.cell(row=r, column=2).value or "",
                        "plan": per, "first": first, "paid": p,
                        "debtY": dy, "debtN": dn,
                        # base - общая часть, own - адресная; share - то, что
                        # ребёнок в этом сборе стоил классу, сумма обеих.
                        "base": share, "own": own, "share": sh, "rest": p - sh})
    spends = [{"date": dt(rs.cell(row=e, column=2).value), "what": rs.cell(row=e, column=4).value or "",
               "amount": num(rs.cell(row=e, column=6).value),
               "proof": str(rs.cell(row=e, column=9).value or "со слов").lower()}
              for e in range(RS_FIRST, RS_LAST + 1)
              if rs.cell(row=e, column=3).value == title and rs.cell(row=e, column=6).value]
    sbory.append({"code": sb.cell(row=r, column=2).value or "", "title": title, "per": per,
                  "first": first, "due": dt(sb.cell(row=r, column=6).value),
                  "dueFull": dt(sb.cell(row=r, column=7).value),
                  "event": dt(sb.cell(row=r, column=8).value),
                  "n": num(sb.cell(row=r, column=9).value), "plan": num(sb.cell(row=r, column=10).value),
                  "collected": num(sb.cell(row=r, column=11).value),
                  "debtY": num(sb.cell(row=r, column=12).value),
                  "debtN": num(sb.cell(row=r, column=13).value),
                  "spent": num(sb.cell(row=r, column=14).value),
                  "rest": num(sb.cell(row=r, column=16).value), "parts": parts, "spends": spends,
                  "sched": schedule(title)})

# Личное списание, привязанное к сбору, в котором ребёнок не участвует (или к сбору
# с другим названием), не попало бы ни в одну строку таблицы: доля ребёнка в шапке
# карточки была бы одна, а расшифровка под ней - другая. Такое молча не проходит.
for k in kids:
    placed = sum(b["own"] for b in k["by"])
    if abs(placed - k["ownSum"]) > 0.005:
        lost = sorted({o["sbor"] or "(сбор не указан)" for o in k["own"]}
                      - {b["sbor"] for b in k["by"]})
        raise SystemExit(
            f"СБОРКА ОСТАНОВЛЕНА: личные списания «{k['raw']}» не разложились по сборам "
            f"({placed:.2f} из {k['ownSum']:.2f}).\n"
            f"Проверьте колонку C листа «Личные списания»: {', '.join(lost) or 'название сбора'} - "
            "такого сбора у ребёнка нет (лист «Участие»).")
    # Доля из журнала (колонка I «Учеников») уже включает личные списания. Если она
    # разошлась с суммой долей по сборам, на странице встретятся два разных ответа
    # на один вопрос - в шапке карточки одно число, в таблице под ней другое.
    if abs(sum(b["share"] for b in k["by"]) - k["share"]) > 0.005:
        raise SystemExit(
            f"СБОРКА ОСТАНОВЛЕНА: у «{k['raw']}» доля расходов из «Учеников» "
            f"({k['share']:.2f}) не сходится с суммой долей по сборам "
            f"({sum(b['share'] for b in k['by']):.2f}).\n"
            "Колонка I «Учеников» обязана равняться сумме «расход на участника» по его "
            "сборам плюс его личные списания.")

# Колонка I - комментарий к взносу: откуда взялась запись. У большинства это
# «из паровозика в родительском чате», но бывает платёж личным сообщением без чека,
# и родителю важно видеть, на чём держится строка в его отчёте. Колонка H
# («Подтверждение (файл)») наружу не идёт: там имя файла с фамилией ребёнка.
for k in kids:
    k["pays"] = [{"date": dt(vz.cell(row=vr, column=2).value),
                  "sbor": vz.cell(row=vr, column=4).value or "",
                  "amount": num(vz.cell(row=vr, column=5).value),
                  "way": vz.cell(row=vr, column=6).value or "",
                  "note": vz.cell(row=vr, column=9).value or ""}
                 for vr in range(VZ_FIRST, VZ_LAST + 1)
                 if vz.cell(row=vr, column=3).value == k["raw"] and vz.cell(row=vr, column=5).value]

# Колонка J - комментарий к расходу: расшифровка чека, «из чего сложилась сумма».
# Родителям он отвечает на «куда ушли деньги» лучше категории, поэтому идёт на
# страницу. Комментарий к сбору (лист «Сборы», колонка Q) - заметка казначея
# и на страницу не выводится.
all_spends = [{"date": dt(rs.cell(row=e, column=2).value), "sbor": rs.cell(row=e, column=3).value or "не привязан",
               "what": rs.cell(row=e, column=4).value or "", "cat": rs.cell(row=e, column=5).value or "без направления",
               "amount": num(rs.cell(row=e, column=6).value),
               "proof": str(rs.cell(row=e, column=9).value or "со слов").lower(),
               "comment": rs.cell(row=e, column=10).value or ""}
              for e in range(RS_FIRST, RS_LAST + 1) if rs.cell(row=e, column=6).value]

EVENTS = []
for s in sbory:
    if s["sched"]:
        # В календарь идёт каждый шаг графика со своей суммой, а не только первый взнос.
        # Отдельного события «полная сумма» больше нет: последний платёж - это просто
        # последний шаг графика, а не самостоятельный показатель.
        for step in s["sched"]:
            EVENTS.append({"date": step["date"], "kind": "due", "code": s["code"],
                           "title": f'Внести {money(step["amount"])} \u20bd - {s["title"]}'})
    elif s["due"] and s["first"]:
        # запасной путь: сбора нет на листе «График». Сбор без объявленной суммы
        # срока не имеет - иначе в календаре появляется «Внести 0 ₽».
        EVENTS.append({"date": s["due"], "kind": "due", "code": s["code"],
                       "title": f'Внести {money(s["first"])} \u20bd - {s["title"]}'})
    if s["event"]:
        EVENTS.append({"date": s["event"], "kind": "event", "code": s["code"],
                       "title": s["title"]})
# Планируемые события идут в тот же календарь, но своей меткой: «срок» - это
# деньги в кассу, «платят сами» - деньги мимо неё. На расчёты не влияют.
EVENTS += plans
DATA = json.dumps({"t": totals, "events": EVENTS, "sbory": sbory,
                   "spends": all_spends, "promises": promises, "flow": flow,
                   "kids": [{a: b for a, b in k.items() if a not in ("row", "col", "raw")} for k in kids]},
                  ensure_ascii=False)

PAGE = r"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex, nofollow, noarchive">
<meta name="referrer" content="no-referrer">
<title>Касса класса 2В - отчёт на __ASOF__</title>
<!-- Пути относительные и без ведущего слэша: сайт живёт в подпапке проекта
     (olegelis.github.io/school-kassa-dashboard/), и «/assets/...» уехал бы
     в корень домена, где ничего нет. -->
<link rel="icon" href="assets/favicon.svg" type="image/svg+xml">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<link rel="manifest" href="manifest.json">
<meta name="apple-mobile-web-app-title" content="Касса 2В">
<meta name="theme-color" content="#2f5496">
<style>
 :root{--ink:#16181d;--dim:#6b7280;--line:#e6e8ec;--bg:#fff;--accent:#2f5496;--good:#0f766e;
       --bad:#b3261e;--warn:#8a6100;--soft:#f7f8fa;--track:#eef1f6;}
 *{box-sizing:border-box;-webkit-tap-highlight-color:transparent;}
 body{margin:0;background:var(--soft);color:var(--ink);
      font:16px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;}
 .wrap{max-width:780px;margin:0 auto;padding:22px 14px 60px;}
 #gate{min-height:100vh;min-height:100dvh;display:flex;align-items:center;justify-content:center;
       padding:20px;}
 .gcard{background:var(--bg);border:1px solid var(--line);border-radius:14px;padding:26px 24px;
        max-width:380px;width:100%;box-shadow:0 4px 24px rgba(22,24,29,.07);}
 /* Без этого блока человек видит голую форму и не понимает, куда попал. */
 .gbrand{display:flex;align-items:center;gap:11px;margin-bottom:16px;}
 .gname{font-size:17px;font-weight:700;letter-spacing:-.01em;min-width:0;}
 .gh{margin:0 0 6px;font-size:19px;} .gp{margin:0 0 16px;color:var(--dim);font-size:14px;}
 #gform{display:flex;gap:8px;} #gform.busy{opacity:.5;pointer-events:none;}
 /* 16px — иначе Safari на iOS зумит страницу при фокусе в поле */
 #gpass{flex:1;min-width:0;padding:11px 13px;border:1px solid var(--line);border-radius:9px;
        font:inherit;font-size:16px;min-height:44px;}
 #gform button{border:0;background:var(--accent);color:#fff;font:inherit;font-weight:600;
               padding:11px 18px;border-radius:9px;cursor:pointer;min-height:44px;}
 .gerr{color:var(--bad);font-size:13.5px;margin-top:10px;min-height:18px;}
 html{scroll-behavior:smooth;}
 /* Резерв под полосу прокрутки нужен только мыши: он держит контент на месте при
    смене вкладок. На тач-экранах scrollbar-gutter не поддерживается, а overflow-y
    на html ломает инерционную прокрутку и даёт второй скролл-контейнер. */
 @media (hover:hover) and (pointer:fine){
  html{scrollbar-gutter:stable;overflow-y:scroll;}
 }
 header{padding:6px 0 18px;border-bottom:2px solid var(--accent);margin-bottom:18px;}
 /* Логотип стоит рядом с блоком «заголовок + подзаголовок», а не рядом с одним
    заголовком: иначе на узком экране подзаголовок переносится под логотип. */
 .brand{display:flex;align-items:center;gap:12px;}
 .logo{flex:none;width:40px;height:40px;display:block;}
 .brand-t{min-width:0;}
 h1{font-size:23px;margin:0 0 5px;letter-spacing:-.01em;}
 .sub{color:var(--dim);font-size:13.5px;}
 .tiles{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:18px;}
 .tiles.c3{grid-template-columns:repeat(3,1fr);}
 .tile{background:var(--bg);border:1px solid var(--line);border-radius:11px;padding:11px 12px;}
 .tile .lab{font-size:10.5px;color:var(--dim);text-transform:uppercase;letter-spacing:.05em;}
 .tile .val{font-size:18px;font-weight:700;margin-top:4px;white-space:nowrap;}
 .tile .tnote{font-size:11.5px;color:var(--dim);margin-top:4px;}
 .tile.rest{border-color:var(--good);} .tile.rest .val{color:var(--good);}
 .tile.owed{border-color:#e3c26b;} .tile.owed .val{color:var(--warn);}
 /* Свободный остаток идёт отдельной строкой во всю ширину: он подводит итог по
    плиткам выше, и только так под цифрой помещается подпись, из чего он получен. */
 .tile.free{grid-column:1/-1;border-color:var(--accent);}
 .tile.free .val{color:var(--accent);}
 /* nowrap: четыре вкладки при 375px переносились во вторую строку и съедали пол-экрана.
    Если подписи всё же не влезают - панель прокручивается вбок, а не растёт вверх. */
 nav{display:flex;flex-wrap:nowrap;overflow-x:auto;gap:6px;background:var(--bg);
     border:1px solid var(--line);
     border-radius:11px;padding:5px;margin-bottom:16px;position:sticky;top:8px;z-index:5;
     box-shadow:0 2px 10px rgba(22,24,29,.06);
     scrollbar-width:none;-ms-overflow-style:none;overscroll-behavior-x:contain;}
 nav::-webkit-scrollbar{display:none;}
 /* С пятой вкладкой полоса перестала помещаться в строку на телефоне. Прокрутка
    вбок здесь была всегда, но до сих пор ни разу не срабатывала, и обрезанная
    вкладка читалась бы как обрезанная, а не как «листайте». Тень у края - тот же
    приём, что у .scroller: background-attachment:local гасит её с той стороны,
    где прокручивать уже нечего. Нарастить полосу на вторую строку нельзя - она
    липкая, и на телефоне это отняло бы у списка ещё сорок точек по высоте. */
 nav{background:
   linear-gradient(to right,var(--bg) 30%,rgba(255,255,255,0)) 0 0,
   linear-gradient(to left,var(--bg) 30%,rgba(255,255,255,0)) 100% 0,
   radial-gradient(farthest-side at 0 50%,rgba(22,24,29,.13),rgba(22,24,29,0)) 0 0,
   radial-gradient(farthest-side at 100% 50%,rgba(22,24,29,.13),rgba(22,24,29,0)) 100% 0,
   var(--bg);
  background-repeat:no-repeat;
  background-size:28px 100%,28px 100%,11px 100%,11px 100%,auto;
  background-attachment:local,local,scroll,scroll,scroll;}
 nav button{flex:1 0 auto;border:0;background:transparent;font:inherit;font-size:14.5px;padding:9px 6px;
            border-radius:8px;cursor:pointer;color:var(--dim);white-space:nowrap;min-height:44px;}
 @media(max-width:400px){
  nav{gap:3px;padding:4px;} nav button{font-size:13.5px;padding:9px 3px;}
 }
 nav button[aria-selected=true]{background:var(--accent);color:#fff;font-weight:600;}
 .card{background:var(--bg);border:1px solid var(--line);border-radius:12px;padding:13px 15px;
       margin-bottom:9px;}
 .card-top{display:flex;align-items:flex-start;gap:10px;}
 .code.sm{font-size:10px;padding:2px 5px;}
 .code{flex:none;font-size:11px;font-weight:700;color:var(--accent);background:#eef2fa;
       border-radius:6px;padding:3px 7px;margin-top:2px;}
 h3{font-size:16px;margin:0;} .meta{color:var(--dim);font-size:12.5px;margin-top:3px;}
 /* График платежей. Разделитель «·» лежит ВНУТРИ шага и попадает под его nowrap,
    поэтому при переносе точка уезжает вместе со своим шагом и не начинает строку,
    а сам шаг не разрывается посередине. */
 .sched{display:flex;flex-wrap:wrap;align-items:baseline;margin:11px 0 2px;font-size:13px;
        line-height:1.7;}
 .sched .st{white-space:nowrap;color:var(--ink);}
 .sched .st:not(:last-child)::after{content:" · ";color:var(--line);white-space:pre;}
 .sched .st b{font-variant-numeric:tabular-nums;font-weight:600;}
 .sched .st.past{color:#a7aeb9;}
 .sched .st.now{color:var(--warn);font-weight:700;}
 .bar{height:7px;background:var(--track);border-radius:99px;margin:11px 0 5px;overflow:hidden;}
 .bar i{display:block;height:100%;background:var(--accent);border-radius:99px;}
 .barlab{display:flex;justify-content:space-between;font-size:12.5px;color:var(--dim);}
 .stats{display:flex;flex-wrap:wrap;gap:6px 16px;margin:10px 0 12px;font-size:13px;
        color:var(--dim);}
 .stats b{color:var(--ink);font-size:14px;font-variant-numeric:tabular-nums;white-space:nowrap;}
 .stats b.good{color:var(--good);} .stats b.bad{color:var(--bad);}
 .acts{display:flex;gap:8px;flex-wrap:wrap;margin-top:4px;}
 /* min-height:44px здесь и у чипов - минимальная цель для пальца (Apple HIG). */
 .btn{border:1px solid var(--line);background:var(--bg);border-radius:8px;padding:7px 12px;
      font:inherit;font-size:13.5px;color:var(--accent);cursor:pointer;
      display:inline-flex;align-items:center;min-height:44px;text-align:left;}
 .btn:hover{background:#f3f6fc;} .btn[aria-expanded=true]{background:#eef2fa;font-weight:600;}
 .chev{display:inline-block;transition:transform .15s;margin-left:5px;font-size:11px;flex:none;}
 .btn[aria-expanded=true] .chev{transform:rotate(180deg);}
 .panel{display:none;margin-top:10px;} .panel.open{display:block;}
 ul.list{list-style:none;margin:0;padding:0;font-size:14.5px;}
 ul.list li{display:flex;align-items:center;gap:9px;padding:8px 0;border-bottom:1px solid var(--line);}
 ul.list li:last-child{border-bottom:none;}
 .dot{flex:none;width:8px;height:8px;border-radius:99px;background:var(--good);}
 .idx{flex:none;min-width:20px;font-size:12px;color:var(--dim);
      font-variant-numeric:tabular-nums;text-align:right;}
 .dot.part{background:#d4a017;} .dot.no{background:var(--bad);}
 .amt{margin-left:auto;font-variant-numeric:tabular-nums;white-space:nowrap;font-weight:600;}
 .amt.bad{color:var(--bad);} .amt.ok{color:var(--good);}
 .amt.soft{color:var(--dim);font-weight:400;}
 /* Возмещение - расход со знаком минус: деньги вернулись в кассу. Красный здесь
    читался бы как долг или ошибка, поэтому цвет тот же, что у остатка. */
 .amt.back{color:var(--good);} .sub b.back{color:var(--good);font-weight:600;}
 .tag{font-size:10.5px;color:var(--dim);border:1px solid var(--line);border-radius:20px;
      padding:1px 7px;white-space:nowrap;}
 /* 16px по той же причине, что и #gpass: меньше - и Safari зумит при фокусе. */
 .search{width:100%;padding:11px 13px;border:1px solid var(--line);border-radius:10px;font:inherit;
         font-size:16px;margin-bottom:10px;background:var(--bg);min-height:44px;}
 .kid{background:var(--bg);border:1px solid var(--line);border-radius:11px;margin-bottom:7px;}
 .kid-h{display:flex;align-items:center;gap:10px;padding:12px 14px;cursor:pointer;}
 .kid-h .nm{font-weight:600;font-size:15px;}
 .kid-h .st{margin-left:auto;text-align:right;font-size:12.5px;color:var(--dim);white-space:nowrap;}
 .kid-h .st b{display:block;font-size:15px;color:var(--ink);}
 .kid-body{display:none;padding:0 14px 14px;} .kid.open .kid-body{display:block;}
 /* Тени по краям показывают, что блок ещё прокручивается вбок: background-attachment:local
    сдвигает «крышки» вместе с контентом, поэтому тень видна только с той стороны,
    где есть непрокрученное. Чистый CSS, без скриптов. */
 .scroller{overflow-x:auto;-webkit-overflow-scrolling:touch;
  background:
   linear-gradient(to right,#fbfcfe 30%,rgba(251,252,254,0)) 0 0,
   linear-gradient(to left,#fbfcfe 30%,rgba(251,252,254,0)) 100% 0,
   radial-gradient(farthest-side at 0 50%,rgba(22,24,29,.16),rgba(22,24,29,0)) 0 0,
   radial-gradient(farthest-side at 100% 50%,rgba(22,24,29,.16),rgba(22,24,29,0)) 100% 0;
  background-repeat:no-repeat;
  background-size:36px 100%,36px 100%,14px 100%,14px 100%;
  background-attachment:local,local,scroll,scroll;}
 table{width:100%;min-width:640px;border-collapse:collapse;font-size:13px;}
 thead th{white-space:nowrap;vertical-align:bottom;line-height:1.2;}
 th{text-align:left;font-size:10.5px;text-transform:uppercase;color:var(--dim);font-weight:600;
    padding:7px 8px;border-bottom:1px solid var(--line);}
 th.num,td.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap;}
 td{padding:8px;border-bottom:1px solid var(--line);} tr:last-child td{border-bottom:none;}
 .empty{background:var(--bg);border:1px dashed var(--line);border-radius:11px;padding:16px;
        color:var(--dim);font-size:14.5px;}
 .note{background:var(--bg);border:1px solid var(--line);border-left:3px solid var(--accent);
       border-radius:9px;padding:12px 15px;margin-top:10px;font-size:14px;color:#39414f;}
 .note b{color:var(--ink);}
 h2{font-size:16px;margin:26px 0 10px;color:var(--accent);}
 .chips{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px;align-items:center;}
 .chip{border:1px solid var(--line);background:var(--bg);border-radius:22px;padding:5px 14px;
       font:inherit;font-size:13px;color:var(--dim);cursor:pointer;
       display:inline-flex;align-items:center;min-height:44px;white-space:nowrap;}
 .chip[aria-pressed=true]{background:var(--accent);color:#fff;border-color:var(--accent);font-weight:600;}
 .rows{background:var(--bg);border:1px solid var(--line);border-radius:11px;overflow:hidden;}
 .row{border-bottom:1px solid var(--line);}
 .row:last-child{border-bottom:none;}
 .row-h{display:flex;align-items:center;gap:7px;padding:9px 11px;cursor:pointer;font-size:14.5px;
        min-height:44px;}
 .row-h .nm{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
 /* Ширина фиксированная, а не min-width: при min-width базисом flex берёт содержимое,
    и «долга нет» даёт другой базис, чем «2 000 ₽» - строки расходятся между собой
    и с заголовком. Те же числа продублированы у .sh.due и .sh.bal. */
 /* Две строки: сверху долг за год, снизу мелким - сколько из него к сроку.
    Ширина выросла с 72px под вторую строку; .sh.due обязан идти следом. */
 .row-h .due{flex:none;width:84px;font-variant-numeric:tabular-nums;
             font-size:13px;text-align:right;line-height:1.25;}
 .row-h .due b{display:block;font-weight:700;color:var(--ink);}
 /* Красным - только те, у кого ближайший платёж ещё не закрыт. У остальных долг
    за год такой же настоящий, но несрочный, и красным он звал бы платить сегодня. */
 .row-h .due.now b{color:var(--bad);}
 .row-h .due i{display:block;font-style:normal;font-size:10px;color:var(--dim);}
 .row-h .due i.free{color:var(--good);}
 .row-h .bal{flex:none;width:84px;padding-left:6px;font-weight:700;text-align:right;
             font-variant-numeric:tabular-nums;white-space:nowrap;}
 .row-h .bal.neg{color:var(--bad);}
 /* Обязательство: в шапке строки две строчки - «на что» и «кому», поэтому это не
    .nm с обрезкой в одну строку. Строка без комментария не раскрывается. */
 .row-h .pm{flex:1;min-width:0;}
 .row-h .pm .sub{display:block;}
 .row.nc .row-h{cursor:default;}
 .row-b{display:none;padding:2px 10px 12px;} .row.open .row-b{display:block;}
 /* gap и padding совпадают с .row-h, иначе заголовки не встают над своими колонками */
 /* letter-spacing живёт на .sh, а не здесь: у контейнера он растягивал скрытый
    шеврон на полпикселя, и правый край заголовков уезжал относительно строк. */
 .row-head{display:flex;align-items:center;gap:7px;padding:0 11px;background:#fbfcfe;
           border-bottom:1px solid var(--line);font-size:10.5px;text-transform:uppercase;
           color:var(--dim);font-weight:600;}
 /* Заголовки переносятся, а не режутся: «Должен за год» и «Не потрачено» в одну
    строку на 390px не помещаются и слипались друг с другом. */
 .sh{border:0;background:transparent;font:inherit;font-size:10.5px;text-transform:uppercase;
     letter-spacing:.04em;color:var(--dim);font-weight:600;cursor:pointer;padding:0;
     display:flex;align-items:center;gap:3px;min-height:44px;line-height:1.25;}
 .sh.due,.sh.bal{flex-wrap:wrap;align-content:center;}
 .sh[aria-pressed=true]{color:var(--accent);}
 /* Стрелка занимает место всегда, даже пустая: иначе колонки дёргаются при смене сортировки */
 .sh .ar{display:inline-block;min-width:9px;font-size:9px;line-height:1;}
 /* flex:1 1 0, как у .row-h .nm: при разном flex-basis свободное место делится
    по-разному и правый край заголовка расходится со строкой на пиксель. */
 .sh.nm{flex:1 1 0;min-width:0;}
 .sh.due{flex:none;width:84px;justify-content:flex-end;text-align:right;}
 .sh.bal{flex:none;width:84px;padding-left:6px;justify-content:flex-end;}
 .pays{margin-top:10px;border-top:1px solid var(--line);padding-top:8px;}
 .pays-h{font-size:10.5px;text-transform:uppercase;letter-spacing:.04em;color:var(--dim);
         font-weight:700;margin-bottom:4px;}
 .pay{display:flex;align-items:baseline;gap:8px;font-size:12.5px;padding:3px 0;}
 .pay .pd{flex:none;font-variant-numeric:tabular-nums;font-weight:600;}
 .pay .pn{color:var(--dim);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
 .pay .pa{margin-left:auto;font-weight:700;font-variant-numeric:tabular-nums;white-space:nowrap;}
 /* Комментарий к платежу - откуда взялась запись. Стоит под своей строкой, а не
    раскрывается по нажатию: платежей у ребёнка единицы, и лишний слой прятал бы
    ровно то, ради чего комментарий заводят - платёж без чека. */
 .pcm{font-size:11.5px;color:var(--dim);line-height:1.4;padding:0 0 5px;}
 /* Итог расшифровки доли расходов: та же цифра, что в шапке разбивки выше,
    поэтому она отбита чертой и не бледная - две одинаковые суммы обязаны
    читаться как одна и та же, а не как ещё одно число. */
 .pay.tot{border-top:1px solid var(--line);margin-top:3px;padding-top:5px;}
 .pay.tot .pn{color:var(--ink);font-weight:600;}
 /* «За что» переносится, а не обрезается многоточием, как название сбора в
    платежах: в платежах сбор один на всю строку и угадывается с первых слов,
    а здесь это единственное место, где написано, что именно ребёнку досталось. */
 .own .pay .pn{white-space:normal;overflow:visible;}
 /* Расшифровка в раскрытой строке: подпись слева, число справа. Раньше здесь
    шла одна строка через «·», но в ней стало пять величин, и главная терялась. */
 .brk{font-size:13px;padding:4px 2px 2px;}
 .brk>div{display:flex;align-items:baseline;gap:12px;padding:6px 0;
          border-bottom:1px dashed var(--line);}
 .brk>div:last-child{border-bottom:none;}
 .brk span{color:var(--dim);}
 .brk b{margin-left:auto;font-variant-numeric:tabular-nums;white-space:nowrap;}
 .brk b.bad{color:var(--bad);}
 .row-h .chev{color:var(--dim);}
 .row.open .row-h .chev{transform:rotate(180deg);}
 .row.open{border:2px solid var(--accent);border-radius:10px;margin:6px 0;
           box-shadow:0 2px 10px rgba(47,84,150,.10);}
 .row.open .row-h{background:#eef2fa;border-radius:8px 8px 0 0;}
 .row.open .row-b{background:#fbfcfe;border-radius:0 0 8px 8px;}
 .hero{background:var(--bg);border:1px solid #e3c26b;border-radius:11px;padding:12px 14px;
        margin-bottom:14px;}
 .hero .lab{display:block;font-size:10.5px;color:var(--warn);text-transform:uppercase;
            letter-spacing:.05em;font-weight:700;margin-bottom:6px;}
 /* Карточка «Ближайшего месяца» - текстовый блок, а не флекс-строка. Во флексе
    заголовок был отдельной колонкой между меткой и отсчётом: длинный
    («Фотографирование класса», «Внести 1 000 ₽ - Учебный год 2026/2027») не
    помещался, уезжал на свою строку целиком и уводил за собой отсчёт - на 390 px
    карточка планируемого события занимала четыре строки против одной у дня
    рождения, и список терял ритм. В блоке заголовок переносится по словам во всю
    ширину карточки, и ни одна запись не выходит за две строки. */
 .hero-i{padding:4px 0;font-size:14.5px;}
 .hero-i b{white-space:nowrap;}
 /* Отсчёт прижат вправо плавающим блоком - во флексе это делал margin-left:auto.
    padding-top выравнивает его по строке: плавающий блок встаёт по верху строки,
    а не по базовой линии, и более мелкий шрифт иначе висит выше заголовка. */
 .hero-i .in{float:right;margin-left:8px;padding-top:3px;
             font-size:12.5px;color:var(--warn);}
 .asof{margin:10px 2px 16px;font-size:12.5px;line-height:1.45;color:var(--dim);}
 .asof b{color:var(--ink);white-space:nowrap;}
 .hero.none{border-color:var(--line);color:var(--dim);font-size:14px;}
 /* Пояснение к событию мимо кассы идёт своей строкой под заголовком: в клетке
    месяца текст обрезан по ширине, и «Ближайший месяц» - единственное место,
    где он читается без наведения мышью (на телефоне его нет). */
 .hero-i .pn{display:block;font-style:normal;font-size:12.5px;color:var(--dim);
             line-height:1.4;margin:1px 0 2px;}
 /* Комментарий из журнала бывает в пять строк - свёрнутый он занимает одну,
    многоточие от line-clamp служит и подсказкой, что текст продолжается.
    Нажатие снимает класс и показывает его целиком: тултипа на телефоне нет,
    а это единственное место, где комментарий читается. */
 .hero-i.clip .pn{display:-webkit-box;-webkit-box-orient:vertical;
                  -webkit-line-clamp:1;line-clamp:1;overflow:hidden;}
 .hero-i.more{cursor:pointer;}
 /* Цена с человека - в начале той же строки, что пояснение, и тёмная: это то,
    что родителю нужно знать до текста, и в свёрнутой строке она видна первой. */
 .hero-i .pn b{color:var(--ink);font-weight:600;white-space:nowrap;}
 /* minmax(0,1fr), а не 1fr: у 1fr минимум - min-content, и длинное название события
    («Полная сумма 5000 ₽ - Учебный год 2026/2027») растягивало колонку шире экрана. */
 .cal{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px;}
 @media(min-width:560px){.cal{grid-template-columns:repeat(3,minmax(0,1fr));}}
 .m{background:var(--bg);border:1px solid var(--line);border-radius:10px;padding:9px 11px 10px;}
 .m.cur{border-color:var(--accent);box-shadow:0 0 0 1px var(--accent) inset;}
 .m h4{margin:0 0 5px;font-size:11px;text-transform:uppercase;letter-spacing:.04em;
       color:var(--accent);font-weight:700;}
 .m.mt h4{color:#b6bcc7;}
 .m ul{list-style:none;margin:0;padding:0;font-size:12.5px;}
 .m li{display:flex;gap:6px;padding:2px 0;align-items:baseline;}
 .m li b{flex:none;min-width:17px;font-variant-numeric:tabular-nums;color:var(--dim);font-weight:600;}
 /* min-width:0 - иначе flex-элемент не сжимается и ellipsis не срабатывает. */
 .m li span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0;}
 .m li.soon b,.m li.soon span{color:var(--warn);font-weight:700;}
 .m li i{font-style:normal;color:var(--dim);font-size:11px;}
 .m li.l-due b,.m li.l-due span{color:var(--bad);} .m li.l-due{font-weight:600;}
 .m li.l-event b,.m li.l-event span{color:var(--accent);}
 /* Событие мимо кассы - зелёным: красный уже занят сроком платежа, синий -
    мероприятием сбора, и третьей денежной строке нельзя читаться как первые две. */
 .m li.l-plan b,.m li.l-plan span{color:var(--good);}
 .k{font-size:10px;text-transform:uppercase;letter-spacing:.03em;border-radius:20px;
    padding:1px 7px;border:1px solid var(--line);color:var(--dim);}
 .k-due{color:var(--bad);border-color:#f3c9c6;} .k-event{color:var(--accent);border-color:#c9d6ee;}
 .k-plan{color:var(--good);border-color:#bcdcd8;}
 .m .none{color:#c9ced7;font-size:12.5px;}
 .mgroup{margin-bottom:12px;}
 .mhead{display:flex;align-items:center;gap:8px;font-size:13px;font-weight:700;
        color:var(--accent);margin:0 0 6px 2px;text-transform:uppercase;letter-spacing:.03em;}
 .mhead span{font-weight:400;color:var(--dim);margin-left:auto;text-transform:none;
              font-variant-numeric:tabular-nums;}
 /* Итог группы уходит в минус, если возмещение перекрыло траты этого дня. */
 .mhead span.back{color:var(--good);font-weight:600;}
 ul.list.pad12{padding:0 12px;} .sub{color:var(--dim);font-size:12px;}
 /* --- Лента кассы («Пришло и ушло») ------------------------------------------
    Таблицы здесь нет намеренно. Пять колонок - дата, операция, пришло, ушло,
    остаток - на 390 px не встают: одна операция бывает в полсотни знаков.
    Табличную вёрстку пришлось бы прокручивать вбок, а прокрутка страницы вбок
    на телефоне ломает и сам список. Поэтому строка сложена в два уровня: сверху
    дата, операция и сумма, снизу справа - сквозной остаток. Остаток при этом
    остаётся ровной колонкой у правого края, а ради неё вкладка и заведена. */
 .mvm{margin-bottom:10px;}
 .mvh{display:grid;grid-template-columns:1fr auto auto;gap:1px 8px;width:100%;
      align-items:baseline;}
 .mvh .mn{font-size:14px;font-weight:700;}
 .mvh .mb{font-variant-numeric:tabular-nums;white-space:nowrap;color:var(--ink);
          font-weight:600;font-size:13px;}
 /* Вторая строка шапки месяца - под всеми тремя колонками первой. */
 .mvh .ms{grid-column:1/-1;font-size:11.5px;color:var(--dim);font-weight:400;
          font-variant-numeric:tabular-nums;}
 .mv{display:flex;flex-wrap:wrap;align-items:baseline;gap:2px 8px;
     padding:8px 11px;border-bottom:1px solid var(--line);font-size:13.5px;}
 .mv:last-child{border-bottom:none;}
 /* Приход и расход различаются подложкой и знаком, а не цветом: зелёное против
    красного читалось бы как «хорошо» против «плохо», а это просто движение -
    деньги пришли или ушли, и ни то ни другое не оценка. */
 .mv.plus{background:#f6f8fc;}
 .mv .d{flex:none;width:36px;font-variant-numeric:tabular-nums;color:var(--dim);
        font-size:12px;}
 /* flex:1 1 0, а не auto: с естественным базисом длинная операция уводила бы
    сумму на свою строку - как это было с заголовком в «Ближайшем месяце». */
 .mv .op{flex:1 1 0;min-width:0;}
 .mv .am{flex:none;font-variant-numeric:tabular-nums;white-space:nowrap;font-weight:600;}
 .mv .bal{flex:0 0 100%;text-align:right;font-size:11.5px;color:var(--dim);
          font-variant-numeric:tabular-nums;}
 .mv .bal i{font-style:normal;}
 /* Шапка колонок появляется только там, где строка укладывается в одну линию и
    заголовкам есть над чем стоять. На телефоне строка двухуровневая, и шапка
    над ней врала бы: колонки там не колонки. Подписывают себя сами - знак «+»
    или «−» у суммы и слово «осталось» у остатка. */
 .mvhd{display:none;}
 @media(min-width:560px){
  .mv .am{flex:none;width:118px;text-align:right;}
  .mv .bal{flex:none;width:152px;font-size:12.5px;}
  /* Со включённой шапкой слово «осталось» в каждой строке повторяет заголовок
     колонки. На телефоне шапки нет, и там оно и подписывает число. */
  .mv .bal i{display:none;}
  .mvhd{display:flex;background:#fbfcfe;font-size:10.5px;text-transform:uppercase;
        letter-spacing:.04em;color:var(--dim);font-weight:600;padding:6px 11px;}
  .mvhd .am,.mvhd .bal{font-weight:600;color:var(--dim);font-size:10.5px;}
 }
 .bd{display:flex;align-items:center;gap:10px;padding:9px 12px;border-bottom:1px solid var(--line);
     font-size:14.5px;}
 .bd:last-child{border-bottom:none;}
 .bd .date{flex:none;min-width:52px;font-weight:700;font-variant-numeric:tabular-nums;}
 .bd .in{margin-left:auto;font-size:12.5px;color:var(--dim);white-space:nowrap;}
 .bd.soon{background:#fff8e6;} .bd.soon .in{color:var(--warn);font-weight:700;}
 .hint{margin:8px 2px 0;font-size:12.5px;color:var(--dim);line-height:1.45;}
 .hint b{color:var(--ink);font-weight:600;}
 .expl{margin-top:18px;}
 .btn.wide{width:100%;text-align:left;padding:10px 14px;display:flex;justify-content:space-between;}
 .expl dl{margin:10px 0 0;background:var(--bg);border:1px solid var(--line);border-radius:11px;
          padding:4px 15px 12px;font-size:13.5px;}
 .expl dt{font-weight:700;margin-top:11px;} .expl dd{margin:2px 0 0;color:#4b5563;}
 footer{margin-top:30px;padding-top:14px;border-top:1px solid var(--line);color:var(--dim);
        font-size:12.5px;}
 @media(max-width:560px){
  .tiles,.tiles.c3{grid-template-columns:1fr 1fr;}
  /* Имя + «внёс …» + «до 30.09 - …» в одну строку при 375px не помещаются:
     разрешаем перенос, сумма остаётся прижатой вправо на второй строке. */
  ul.list li{flex-wrap:wrap;row-gap:2px;}
  /* Полное «Фамилия Имя» рядом с двумя колонками чисел на 390px не умещается,
     а обрезать его многоточием - терять ровно то, ради чего имя сделали полным
     («Воробушкина С…» не найти глазами). Разрешаем перенос: выше становятся
     только те строки, которым не хватило места.
     overflow-wrap - на 320px не помещается уже одна «Воробушкина»: без него
     слово вылезло бы на колонку с суммой. */
  .row-h .nm{white-space:normal;overflow-wrap:anywhere;}
 }
 /* Таблица по сборам ребёнка вместо прокрутки на 640px разбирается в карточки
    «показатель - значение»: на телефоне боковой скролл внутри карточки не читается.
    Порог 700, а не 560: таблице нужно 640px плюс поля, и на iPhone SE в альбомной
    ориентации (667) у неё срезало последнюю колонку. */
 @media(max-width:700px){
  .row-b .scroller{overflow-x:visible;background:none;}
  .row-b table{min-width:0;}
  .row-b thead{display:none;}
  .row-b table,.row-b tbody,.row-b tr,.row-b td{display:block;}
  .row-b tr{border:1px solid var(--line);border-radius:10px;margin-bottom:9px;
            background:var(--bg);overflow:hidden;}
  .row-b tr:last-child{margin-bottom:0;}
  .row-b td{display:flex;align-items:baseline;justify-content:space-between;gap:14px;
            padding:8px 11px;border-bottom:1px solid var(--line);text-align:right;}
  .row-b tr td:last-child{border-bottom:none;}
  .row-b td::before{content:attr(data-l);flex:none;text-align:left;color:var(--dim);
                    font-size:10.5px;text-transform:uppercase;letter-spacing:.04em;
                    font-weight:600;line-height:1.6;}
  .row-b td.hd{display:block;background:#eef2fa;font-weight:700;text-align:left;
               font-size:13.5px;padding:9px 11px;}
  .row-b td.hd::before,.row-b td.non::before{content:none;}
  .row-b td.non{display:block;text-align:left;}
 }
 @media print{
  body{background:#fff;} .wrap{max-width:none;padding:0;}
  nav,.acts,.search{display:none!important;}
  [hidden]{display:block!important;}
  .panel{display:block!important;} .kid-body{display:block!important;}
  .card,.kid{break-inside:avoid;page-break-inside:avoid;}
  table{min-width:0;} .scroller{overflow:visible;}
  #pane-kids::before{content:"По каждому ребёнку";display:block;font-size:16px;
   color:var(--accent);font-weight:700;margin:24px 0 10px;}
  #pane-flow::before{content:"Пришло и ушло";display:block;font-size:16px;
                     font-weight:700;margin:0 0 8px;}
  #pane-spends::before{content:"Все траты";display:block;font-size:16px;
   color:var(--accent);font-weight:700;margin:24px 0 10px;}
  .expl .btn{display:none!important;} .expl dl{break-inside:avoid;}
  .row-b{display:block!important;} .row{break-inside:avoid;} .chips{display:none!important;}
 }
</style></head><body>
 <div id="gate">
  <div class="gcard">
   <div class="gbrand">
    <img class="logo" src="assets/logo.svg" width="40" height="40" alt="">
    <div class="gname">Касса класса 2В</div>
   </div>
   <h2 class="gh">Доступ по коду</h2>
   <p class="gp">Отчёт закрыт от посторонних. Введите код, который дала Аня.</p>
   <form id="gform" autocomplete="on">
    <input id="gpass" type="password" name="password" autocomplete="current-password"
           placeholder="код" autocapitalize="off" autocorrect="off" spellcheck="false">
    <button type="submit">Открыть</button>
   </form>
   <div class="gerr" id="gerr"></div>
  </div>
 </div>
<div class="wrap" id="app" hidden>
 <header>
  <div class="brand">
   <img class="logo" src="assets/logo.svg" width="40" height="40" alt="">
   <div class="brand-t"><h1>Касса класса 2В</h1>
    <div class="sub">Учебный год 2026/2027 · отчёт на __ASOF__ · казначей __TREAS__</div></div>
  </div>
 </header>
 <div class="tiles" id="tiles"></div>
 <!-- Дата денег стоит вплотную к плиткам, а не только в подписи под заголовком:
      именно здесь родитель читает остаток и именно здесь может решить, что тот
      живой. Календарь ниже считает дни от настоящего сегодня, поэтому сказано
      прямо, что деньгам это не касается. -->
 <p class="asof">Деньги посчитаны на <b>__ASOF__</b> - это день сборки отчёта.
  Взносы и траты после этой даты в суммы не вошли: чтобы они появились, отчёт
  пересобирают. Календарь событий и дней рождения ниже - живой, он считает дни
  от сегодняшнего дня.</p>
 <span id="navtop"></span>
 <nav role="tablist">
  <button role="tab" data-tab="bdays" aria-selected="true">События</button>
  <button role="tab" data-tab="sbory" aria-selected="false">Сборы</button>
  <button role="tab" data-tab="kids" aria-selected="false">По детям</button>
  <button role="tab" data-tab="spends" aria-selected="false">Расходы</button>
  <button role="tab" data-tab="flow" aria-selected="false">Пришло и ушло</button>
 </nav>
 <div id="pane-bdays"></div>
 <div class="expl" id="ex-bdays"></div>
 <div id="pane-sbory" hidden></div>
 <div class="expl" id="ex-sbory" hidden></div>
 <div id="pane-kids" hidden>
  <input class="search" id="q" type="search" placeholder="__SEARCHPH__" autocomplete="off">
  <div class="chips" id="chips"></div>
  <div id="kidlist"></div>
  <div id="kidext"></div>
  <div class="hint" id="kidhint"></div>
  <div class="expl" id="ex-kids"></div>
 </div>
 <div id="pane-spends" hidden></div>
 <div class="expl" id="ex-spends" hidden></div>
 <div id="pane-flow" hidden></div>
 <div class="expl" id="ex-flow" hidden></div>

 <footer>Сформировано __ASOF__ автоматически из журнала «Касса_2В_2026-2027.xlsx».
  Цифры руками не набираются. Версия __STAMP__.</footer>
 </div>
</div>
<script>
const RAW=__DATA__;

async function decryptPayload(b64, pass){
 // Та же нормализация, что в norm_pass() генератора: регистр не важен, пробелы срезаются.
 // Меняется здесь - меняется и там, иначе ключи разойдутся и страница не откроется.
 const norm=String(pass).trim().toLowerCase();
 const bin=Uint8Array.from(atob(b64), c=>c.charCodeAt(0));
 const salt=bin.slice(0,16), iv=bin.slice(16,28), ct=bin.slice(28);
 const km=await crypto.subtle.importKey('raw',new TextEncoder().encode(norm),'PBKDF2',false,['deriveKey']);
 const key=await crypto.subtle.deriveKey(
   {name:'PBKDF2',salt:salt,iterations:__ITERS__,hash:'SHA-256'},km,
   {name:'AES-GCM',length:256},false,['decrypt']);
 const out=await crypto.subtle.decrypt({name:'AES-GCM',iv:iv},key,ct);
 return JSON.parse(new TextDecoder().decode(out));
}

function boot(D){

// Минус выводится знаком U+2212, а не дефисом: toLocaleString даёт «-600», и на
// мелком кегле дефис теряется рядом с цифрой - сумма читается как обычная трата.
// Настоящий минус по ширине равен цифре и виден сразу.
const rub=n=>{n=Math.round((n||0)*100)/100;const s=Number.isInteger(n)?n.toLocaleString('ru-RU'):
  n.toLocaleString('ru-RU',{minimumFractionDigits:2,maximumFractionDigits:2});
  return s.replace(/ /g,' ').replace(/^-/,'−')+' ₽';};
const esc=s=>String(s==null?'':s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
// Склонение по числу: форма для 1, для 2-4 и для 5 и больше, с исключением на 11-14.
const plural=(n,one,few,many)=>{const a=n%10,b=n%100;
 return b>=11&&b<=14?many:(a===1?one:(a>=2&&a<=4?few:many));};
const T=D.t;
// На странице два разных «сегодня», и путать их нельзя.
//
// ASOF - дата сборки. Всё денежное считается на неё и только на неё: остаток,
// долги, доли, «ждём к сроку», подсветка шагов графика. Живой её делать нельзя -
// страница не знает ни одного платежа, прошедшего после сборки, и назавтра сама
// объявила бы должником того, кто уже заплатил.
//
// NOW - настоящее сегодня в браузере, с обнулённым временем суток. Уходит только
// в календарь: обратный отсчёт до дня рождения, «Ближайший месяц», рамка текущего
// месяца, жёлтая подсветка. Эти вещи от денег не зависят, и ради них страницу
// больше не надо пересобирать.
const ASOF=new Date(__YEAR__,__MONTH__-1,__DAY__), ASOFTXT='__ASOF__';
// Полночь по часам читателя. Отсчёт ведётся между полуночями, поэтому считаются
// календарные дни, а не сутки по 24 часа: в поясе с переводом стрелок разница
// выходит 23 или 25 часов, и Math.round ниже возвращает целое число дней.
const NOW=(()=>{const n=new Date();n.setHours(0,0,0,0);return n;})();
const parseD=s=>{const [d,m,y]=String(s).split('.').map(Number);return new Date(y,m-1,d);};
// Дата ближайшего срока нигде не пишется строкой: планка едет с каждым шагом
// графика, и через месяц вместо 30.09 там 31.10. Берётся due сбора - то самое
// поле, из которого журнал считает и саму сумму «надо к сроку», так что число
// и его дата не могут разойтись.
const dm=d=>String(d||'').slice(0,5);          // «30.09.2026» -> «30.09»
const DUE={};                                   // сбор -> его ближайший срок
const SCHED={};                                 // сбор -> шаги графика платежей
D.sbory.forEach(s=>{if(s.due)DUE[s.title]=s.due;
 if(s.sched&&s.sched.length)SCHED[s.title]=s.sched;});
// Что ребёнку осталось внести по графику сбора. Идём по тем же шагам, что
// нарисованы в карточке сбора, нарастающим итогом и вычитаем уже внесённое.
// Шаги не позже ближайшего срока схлопываются в один, датированный этим сроком:
// ровно так журнал считает «не хватает к сроку», и просрочка на странице не
// заводится - пропущенный платёж от 25.08 показан не отдельной красной датой,
// а частью планки 30.09.
const planLeft=b=>{
 const sc=SCHED[b.sbor]; if(!sc)return null;
 let cum=0; const out=[];
 sc.forEach(x=>{const prev=cum; cum+=x.amount;
  const need=Math.max(0,cum-b.paid)-Math.max(0,prev-b.paid);
  if(need>0.005)out.push({date:x.date,amount:need});});
 const due=DUE[b.sbor];
 if(due&&out.length){
  const head=out.filter(x=>parseD(x.date)<=parseD(due));
  if(head.length)out.splice(0,head.length,
   {date:due,amount:head.reduce((a,x)=>a+x.amount,0)});}
 // Расшифровка обязана сходиться с числом из журнала: две цифры об одном и том же
 // не должны расходиться на глазах у родителя. Не сошлось - расшифровки не будет.
 return Math.abs(out.reduce((a,x)=>a+x.amount,0)-b.debtY)<0.005?out:null;};
// Дата для плитки в шапке. Плитка складывает нехватку по всем сборам, а сроки у
// них могут быть разные - тогда одной даты у суммы нет. Считаются только сборы,
// у которых к сроку чего-то не хватает: сбор без долга в плитку ничего не вносит,
// даже если срок у него есть. Из оставшихся берётся самый ранний - деньги нужны
// первыми к нему; когда таких дат несколько, подпись честно говорит «ближайший».
const dueDates=[...new Set(D.sbory.filter(s=>s.debtN>0&&s.due).map(s=>s.due))]
 .sort((a,b)=>parseD(a)-parseD(b));
// Главное число долга - годовое: родитель, закрывший ближайший платёж, иначе
// читает «к сроку 0» как «ничего не должен». Срок ушёл в подпись под суммой.
// Порядок плиток: три про кассу, потом долг, потом свободный остаток - долг
// родителям важнее того, сколько из кассы ещё никому не обещано.
// Подпись про срок живой не делается: страница не знает платежей, прошедших
// после сборки, и, подняв планку сама, назвала бы должником того, кто заплатил.
// Но и молчать нельзя: календарь рядом живой, и после 30.09 он показывает уже
// 31.10, а плитка - всё ещё «до 30.09». Два ответа на один вопрос в одном экране.
// Поэтому сумма остаётся как есть, а приписка объясняет расхождение: цифра не
// протухла, она просто посчитана на день сборки, и деньги никуда не делись.
const dueNote=(()=>{
 if(!T.dueNow)return 'к ближайшему сроку внесено всё';
 if(!dueDates.length)return `из них ${rub(T.dueNow)}`;
 const s=`из них ${rub(T.dueNow)}`+(dueDates.length>1?' к ближайшему сроку ':' до ')+dm(dueDates[0]);
 return parseD(dueDates[0])<NOW
  ? `${s} - срок прошёл, сумма посчитана на ${ASOFTXT}: платежи после этой даты сюда не попали`
  : s;})();
const tiles=[['Собрано',T.collected,''],['Потрачено',T.spent,''],['Остаток',T.rest,'rest']];
if(T.dueYear)tiles.push(['Осталось внести за год',T.dueYear,'owed',dueNote]);
// «Остаток» - все деньги кассы, «Свободный остаток» - то, что из них ещё никому
// не обещано. Плитка появляется только когда обязательства есть.
if(T.promised)tiles.push(['Свободный остаток',T.free,'free',
 `остаток кассы минус ${rub(T.promised)} уже обещанных`]);
document.getElementById('tiles').innerHTML=tiles.map(([l,v,c,n])=>
 `<div class="tile ${c}"><div class="lab">${l}</div><div class="val">${rub(v)}</div>${
   n?`<div class="tnote">${n}</div>`:''}</div>`).join('');
// Считаем плитки верхнего ряда: широкий «Свободный остаток» в сетку колонок не входит.
if(tiles.length-(T.promised?1:0)===3)document.getElementById('tiles').classList.add('c3');

// Сортировка одна на всю страницу: и список детей, и участники внутри сбора.
// Первый клик по колонке даёт направление из SDEF, повторный - разворачивает.
let SORT='name', SDIR=1;
const SDEF={name:1,debt:-1,rest:1};
// ord - порядок строк в журнале, вторичный ключ: суммы часто совпадают (у многих ровно
// по 2000 к сроку), и без него строки при равных значениях прыгали бы при перерисовке.
const cmp=(a,b)=>{
 let r;
 if(SORT==='debt')r=(a.debtY||0)-(b.debtY||0)||(a.debtN||0)-(b.debtN||0);
 else if(SORT==='rest')r=(a.rest||0)-(b.rest||0);
 else r=String(a.name).localeCompare(String(b.name),'ru');
 return (r*SDIR)||((a.ord||0)-(b.ord||0));};
const sortHead=(key,label,cls)=>
 `<button class="sh ${cls}" data-sort="${key}" aria-pressed="${SORT===key}">${label}<span
   class="ar">${SORT===key?(SDIR===1?'▲':'▼'):''}</span></button>`;

function renderSbory(){
document.getElementById('pane-sbory').innerHTML=D.sbory.map((s,i)=>{
 const pct=s.plan?Math.min(100,Math.round(s.collected/s.plan*100)):0;
 const done=s.parts.filter(p=>!p.debtN).length;
 const parts=[...s.parts].sort(cmp).map((p,pi)=>{
  const cls=!p.debtN?(p.debtY?'':''):(p.paid?'part':'no');
  // Срок сбора, а не общий: в этом списке у всех нехватка по одному и тому же сбору.
  const right=p.debtN?`<span class="amt bad">${s.due?`до ${dm(s.due)} - `:'нужно '}${rub(p.debtN)}</span>`
    :(p.debtY?`<span class="amt soft">ещё ${rub(p.debtY)} за год</span>`
             :`<span class="amt ok">рассчитался</span>`);
  return `<li><span class="dot ${cls}"></span><span class="idx">${pi+1}</span>${esc(p.name)}
    <span class="tag">внёс ${rub(p.paid)}</span>${right}</li>`;}).join('');
 const sp=s.spends.length?s.spends.map(e=>`<li>${e.date?`<span class="tag">${esc(e.date)}</span>`:''}
   ${esc(e.what)}${e.amount<0?' <span class="tag">возмещение</span>':''}
   <span class="amt${e.amount<0?' back':''}">${rub(e.amount)}</span>
   <span class="tag">${esc(e.proof)}</span></li>`).join('')
  :'<li style="color:var(--dim)">Из этого сбора пока ничего не потрачено.</li>';
 // График платежей заменяет обе денежные метки: и «по 2 000 ₽ до 25.08», и
 // «всего 5 000 ₽ до 30.11» - оба факта видны в самих шагах. Без графика
 // (сбора нет на листе «График») метки остаются как раньше.
 const sc=s.sched||[];
 const nextI=sc.findIndex(x=>parseD(x.date)>=ASOF);
 const schedHtml=sc.length?`<div class="sched">${sc.map((x,i)=>{
   const cls=i===nextI?'now':(parseD(x.date)<ASOF?'past':'');
   return `<span class="st ${cls}"><b>${esc(x.date.slice(0,5))}</b> - ${rub(x.amount)}</span>`;
  }).join('')}</div>`:'';
 const meta=[`${s.n} ${plural(s.n,'участник','участника','участников')}`];
 if(s.event)meta.push(`событие ${s.event.slice(0,5)}`);
 if(!sc.length){
  if(s.first)meta.push(`по ${rub(s.first)}${s.due?' до '+s.due.slice(0,5):''}`);
  if(s.per&&s.per!==s.first)meta.push(`всего ${rub(s.per)}${s.dueFull?' до '+s.dueFull.slice(0,5):' - платежами в течение года'}`);
  // Ни графика, ни суммы за год, ни взноса к сроку: сбор объявлен, цифры ещё нет.
  // Прогресс-бар при пустом плане и так не рисуется.
  if(!s.per&&!s.first)meta.push('сумма пока не объявлена');
 }
 return `<section class="card">
  <div class="card-top">${s.code?`<span class="code">${esc(s.code)}</span>`:''}
   <div><h3>${esc(s.title)}</h3><div class="meta">${meta.join(' · ')}</div></div></div>
  ${schedHtml}
  ${s.plan?`<div class="bar"><i style="width:${pct}%"></i></div>
   <div class="barlab"><span>собрано ${rub(s.collected)} из ${rub(s.plan)}</span><span>${pct}%</span></div>`:''}
  <div class="stats">
   <span>потрачено <b>${rub(s.spent)}</b></span>
   <span>остаток <b class="good">${rub(s.rest)}</b></span>
   <!-- Долг за год стоит перед сроком и здесь: иначе вкладка «Сборы» показывала бы
        главным то число, которое на остальной странице стало расшифровкой. -->
   <span>должны за год <b${s.debtY?' class="bad"':''}>${rub(s.debtY)}</b></span>
   <span>из них к сроку${s.due?' '+dm(s.due):''} <b${s.debtN?' class="bad"':''}>${rub(s.debtN)}</b></span>
  </div>
  <div class="acts">
   <button class="btn" data-toggle="pp${i}" aria-expanded="false">Участники · внесли к сроку ${done} из ${s.parts.length}<span class="chev">▾</span></button>
   <button class="btn" data-toggle="ps${i}" aria-expanded="false">Расходы · ${s.spends.length}<span class="chev">▾</span></button>
  </div>
  <div class="panel" id="pp${i}"><ul class="list">${parts}</ul></div>
  <div class="panel" id="ps${i}"><ul class="list">${sp}</ul></div></section>`;}).join('')
 ||'<div class="empty">Сборов пока нет.</div>';}
renderSbory();

// Дети класса и внешние плательщики разведены: OWN - те, кого касаются чипы,
// поиск, сортировка и деление расходов; EXT - чужие деньги, лежащие в той же
// кассе (сосед по кабинету скинулся на декор). Их остаток входит в остаток
// кассы, поэтому он показан отдельным блоком, а не спрятан.
const OWN=D.kids.filter(k=>!k.ext),
      EXT=D.kids.filter(k=>k.ext&&!k.teach&&k.rest!==0);
// Личные списания - часть уже потраченного, адресованная конкретным детям: вещь
// досталась одному, её и оплачивает он, а не весь класс. Поровну поэтому делится
// не всё «потрачено», а остаток за вычетом адресных трат - иначе подпись под
// таблицей называла бы долю, которой на самом деле нет ни у кого.
const PERS=OWN.reduce((a,k)=>a+(k.ownSum||0),0);
let KFILTER='all';
function renderKids(f){f=(f||'').trim().toLowerCase();
 let L=OWN.filter(k=>!f||k.name.toLowerCase().includes(f));
 L=[...L].sort(cmp);
 if(KFILTER==='debt')L=L.filter(k=>k.debtY>0);
 if(KFILTER==='ok')L=L.filter(k=>!k.debtY);
 document.getElementById('kidlist').innerHTML=L.length?`<div class="rows">
  <div class="row-head"><span class="idx"></span>
   ${sortHead('name','Ученик','nm')}${sortHead('debt','Должен за год','due')}${sortHead('rest','Не потрачено','bal')}
   <span class="chev" style="visibility:hidden">▾</span></div>
  ${L.map((k,i)=>{
  // Нехватка ребёнка сложена по всем его сборам, поэтому дата берётся не общая,
  // а самая ранняя из тех сборов, где ему к сроку не хватает.
  const kdue=k.by.filter(b=>b.debtN>0&&DUE[b.sbor]).map(b=>DUE[b.sbor])
   .sort((a,b)=>parseD(a)-parseD(b))[0];
  // Оставшиеся платежи по всем сборам ребёнка, сведённые по датам. Сумма шагов
  // равна «осталось за год», первый шаг - «из них к сроку»: обе цифры из шапки
  // строки видны здесь в разбивке, а не появляются ниоткуда.
  const steps=[];
  k.by.forEach(b=>(planLeft(b)||[]).forEach(st=>{
   const e=steps.find(y=>y.date===st.date);
   if(e)e.amount+=st.amount; else steps.push({date:st.date,amount:st.amount});}));
  steps.sort((a,b)=>parseD(a.date)-parseD(b.date));
  // data-l дублирует заголовок столбца: на узких экранах thead скрыт, и подпись
  // берётся из атрибута через ::before (см. @media max-width:560px).
  const rows=k.by.map(b=>`<tr>
    <td class="hd">${b.code?`<span class="code sm">${esc(b.code)}</span> `:''}${esc(b.sbor)}</td>
    <td class="num" data-l="План к сроку">${rub(b.first)}</td>
    <td class="num" data-l="План за год">${rub(b.plan)}</td>
    <td class="num" data-l="Внесено">${rub(b.paid)}</td>
    <td class="num" data-l="Должен за год"${b.debtY?' style="font-weight:700"':''}>${b.debtY?rub(b.debtY):'-'}</td>
    <td class="num" data-l="Из них к сроку"${b.debtN?' style="color:var(--bad);font-weight:700"':''}>${b.debtN?rub(b.debtN):'-'}</td>
    <td class="num" data-l="Доля расходов">${rub(b.share)}</td>
    <td class="num" data-l="Не потрачено"><b>${rub(b.rest)}</b></td></tr>`).join('')
   ||'<tr><td class="non" colspan="8" style="color:var(--dim)">Ни в одном сборе не участвует.</td></tr>';
  // Номер - позиция в текущем списке, а не закреплённый за ребёнком: пересчитывается
  // при каждой сортировке и фильтрации. data-k держит скрытый ключ для восстановления
  // раскрытых строк после перерисовки.
  return `<div class="row" data-k="${k.ord}"><div class="row-h">
    <span class="idx">${i+1}</span>
    <span class="nm">${esc(k.name)}</span>
    <span class="due${k.debtN?' now':''}">${k.debtY
      ?`<b>${rub(k.debtY)}</b><i>${k.debtN
         ?`из них ${rub(k.debtN)}${kdue?' до '+dm(kdue):''}`
         :'ближайший срок закрыт'}</i>`
      :'<i class="free">долга нет</i>'}</span>
    <span class="bal${k.rest<0?' neg':''}">${rub(k.rest)}</span>
    <span class="chev">▾</span></div>
   <div class="row-b">
    <div class="brk">
     <div><span>Внесено</span><b>${rub(k.paid)}</b></div>
     <div><span>Доля расходов</span><b>${rub(k.share)}</b></div>
     <div><span>Не потрачено из внесённого</span>
      <b${k.rest<0?' class="bad"':''}>${rub(k.rest)}</b></div>
     <div><span>Осталось внести за год</span>
      <b${k.debtY?' class="bad"':''}>${rub(k.debtY)}</b></div>
     ${k.debtY?`<div><span>${k.debtN?`Из них к ближайшему сроку${kdue?', '+dm(kdue):''}`
       :'Ближайший срок'}</span>
      <b${k.debtN?' class="bad"':''}>${k.debtN?rub(k.debtN):'закрыт'}</b></div>`:''}
    </div>
    ${k.ownSum?`<div class="pays own"><div class="pays-h">Из чего сложилась доля расходов</div>
      <div class="pay"><span class="pn">Общая часть, поровну со всеми</span>
       <span class="pa">${rub(k.share-k.ownSum)}</span></div>
      ${k.own.map(o=>`<div class="pay"><span class="pd">${esc(o.date)}</span>
        <span class="pn">${esc(o.what)}</span>
        <span class="pa">${rub(o.amount)}</span>
        <span class="tag">${esc(o.proof)}</span></div>${
        o.note?`<div class="pcm">${esc(o.note)}</div>`:''}`).join('')}
      <div class="pay tot"><span class="pn">Доля расходов</span>
       <span class="pa">${rub(k.share)}</span></div>
      <div class="pcm">Это доля расходов, а не долг: по графику взносов ребёнок
       должен ровно столько же, сколько все остальные.</div>
     </div>`:''}
    ${steps.length?`<div class="pays plan"><div class="pays-h">Осталось внести по графику</div>
      <div class="sched">${steps.map((x,si)=>
       `<span class="st${si===0?' now':''}"><b>${dm(x.date)}</b> - ${rub(x.amount)}</span>`
      ).join('')}</div></div>`:''}
    <div class="scroller"><table>
    <thead><tr>
     <th>Сбор</th>
     <th class="num">План<br>к сроку</th>
     <th class="num">План<br>за год</th>
     <th class="num">Внесено</th>
     <th class="num">Должен<br>за год</th>
     <th class="num">Из них<br>к сроку</th>
     <th class="num">Доля<br>расходов</th>
     <th class="num">Не<br>потрачено</th></tr></thead>
    <tbody>${rows}</tbody></table></div>
    ${k.pays.length?`<div class="pays"><div class="pays-h">Платежи</div>
      ${k.pays.map(pp=>`<div class="pay"><span class="pd">${esc(pp.date)}</span>
        <span class="pn">${esc(pp.sbor)}</span>
        <span class="pa">${rub(pp.amount)}</span>
        <span class="tag">${esc(String(pp.way).toLowerCase())}</span></div>${
        pp.note?`<div class="pcm">${esc(pp.note)}</div>`:''}`).join('')}</div>`
     :'<div class="pays"><div class="pays-h">Платежей ещё не было</div></div>'}
    </div></div>`;}).join('')}</div>`
  :'<div class="empty">Никого не нашли.</div>';
 const h=document.getElementById('kidhint');
 if(h)h.innerHTML=`<b>Должен за год</b> - сколько ещё осталось внести до полной суммы сбора;
  мелким шрифтом под ним - какая часть этой суммы нужна к ближайшему сроку.
  <b>Не потрачено</b> - деньги ребёнка, ещё не ушедшие в расходы: внесено минус своя доля
  (${PERS?`общие расходы ${rub(D.t.spent-PERS)}`:`уже потрачено ${rub(D.t.spent)}`} ÷ ${
   OWN.length} = ${rub((D.t.spent-PERS)/OWN.length)} с человека${PERS
   ?`; сверх этого у некоторых детей есть личные списания - что именно, видно внутри строки`:''}).
  Это не «долга нет»: долг и не потраченные деньги - разные вещи, и бывают вместе.
  Минус означает, что за ребёнка уже потрачено больше, чем он внёс.`;}

// Блок внешних плательщиков не зависит от поиска, сортировки и чипов: он собирается
// один раз и стоит между таблицей и подписью. Без него сумма видимых остатков
// не сходится с остатком кассы.
// Учитель тоже «не ученик», но денег за ним нет - в блоке поступлений ему нечего
// показывать. Отсекается и по признаку, и по нулевой сумме: признак говорит, что
// это за строка, а сумма - что в блоке о ней вообще можно написать.
document.getElementById('kidext').innerHTML=EXT.length?`<h2>Поступления не от учеников класса</h2>
 <div class="rows"><ul class="list pad12">${EXT.map(k=>
  `<li>${esc(k.name)}<span class="amt">${rub(k.rest)}</span></li>`).join('')}</ul></div>`:'';

// Фильтр делит по главному вопросу - должен за год или нет. Прежние «Надо к сроку»
// и «Только за год» делили по сроку, а он теперь вторичен: срочность видна прямо
// в строке («из них 1 000 ₽ до 30.09» либо «ближайший срок закрыт»), и ради неё
// больше не нужно переключать фильтр. Подпись чипа совпадает с плиткой в шапке.
const cnt={all:OWN.length,debt:OWN.filter(k=>k.debtY>0).length,
 ok:OWN.filter(k=>!k.debtY).length};
document.getElementById('chips').innerHTML=[['all','Все'],['debt','Есть долг за год'],
 ['ok','Рассчитались']].map(([k,l])=>
 `<button class="chip" data-chip="${k}" aria-pressed="${k==='all'}">${l} · ${cnt[k]}</button>`).join('');
document.getElementById('chips').addEventListener('click',e=>{
 const c=e.target.closest('.chip'); if(!c)return;
 KFILTER=c.dataset.chip;
 document.querySelectorAll('.chip').forEach(x=>x.setAttribute('aria-pressed',String(x===c)));
 renderKids(document.getElementById('q').value);});
renderKids('');
document.getElementById('q').addEventListener('input',e=>renderKids(e.target.value));

// дни рождения
const bd=D.kids.filter(k=>k.bday).map(k=>{
 const [d,m]=k.bday.split('.').map(Number);
 let y=NOW.getFullYear(), next=new Date(y,m-1,d);
 if(next<NOW)next=new Date(y+1,m-1,d);
 return {name:k.name,role:k.teach,date:k.bday,in:Math.round((next-NOW)/86400000)};});
bd.sort((a,b)=>a.in-b.in);
const MONT=['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август',
 'Сентябрь','Октябрь','Ноябрь','Декабрь'];
const curM=NOW.getMonth();
const daysTo=(d,m)=>{let y=NOW.getFullYear(),n=new Date(y,m-1,d);
 if(n<NOW)n=new Date(y+1,m-1,d);return {in:Math.round((n-NOW)/86400000),y:n.getFullYear()};};
const ALL=[];
bd.forEach(b=>{const [d,m]=b.date.split('.').map(Number);const r=daysTo(d,m);
 // Имя в клетке календаря больше не сокращается: сокращённая форма вернула бы
 // на страницу ровно ту подпись, от которой родком отказался. Длинное имя
 // обрезается многоточием средствами CSS (.m li span), а не логикой.
 ALL.push({kind:b.role?'teacher':'bd',d:d,m:m,in:r.in,name:b.name});});
// События журнала - разовые, и год у них в дате есть. Переносить их на год
// вперёд, как день рождения, нельзя: шаг графика «Внести 2 000 ₽ до 25.08» - это
// первый взнос конкретного сбора, в августе следующего года его уже не будет,
// а красная строка в календаре звала бы платить второй раз. То же у мероприятия
// сбора и у планируемого события: съёмка 08.09.2026 в сентябре 2027 не повторится.
// Поэтому отсчёт идёт по настоящей дате со своим годом, а прошедшее просто
// уходит из календаря - оно уже случилось. Дни рождения, наоборот, ежегодны,
// их перенос на следующий год выше остаётся как был.
(D.events||[]).forEach(e=>{const [d,m,y]=e.date.split('.').map(Number);
 const on=new Date(y,m-1,d); if(on<NOW)return;
 ALL.push({kind:e.kind,d:d,m:m,in:Math.round((on-NOW)/86400000),name:e.title,
  code:e.code,amt:e.amt||'',note:e.note||''});});
ALL.sort((a,b)=>a.in-b.in);
// «срок» - деньги в кассу, «платят сами» - мимо неё, родитель отдаёт их напрямую
// организатору. Подпись «событие» занята мероприятием сбора, оно оплачено кассой.
const KIND={bd:'др',teacher:'др',due:'срок',event:'событие',plan:'платят сами'};
const dd=n=>String(n).padStart(2,'0');
const yr=n=>{const a=n%10,b=n%100;
 if(b>=11&&b<=14)return n+' лет'; if(a===1)return n+' год';
 if(a>=2&&a<=4)return n+' года'; return n+' лет';};
const soonList=ALL.filter(e=>e.in<=30);
const byMonth={};ALL.forEach(e=>{(byMonth[e.m-1]=byMonth[e.m-1]||[]).push(e);});
Object.values(byMonth).forEach(a=>a.sort((x,y)=>x.d-y.d));
const order=[];for(let i=0;i<12;i++)order.push((8+i)%12);
document.getElementById('pane-bdays').innerHTML=ALL.length?`
 ${soonList.length?`<div class="hero"><span class="lab">Ближайший месяц</span>
   ${soonList.map(e=>`<div class="hero-i clip"${
     e.note?` title="${esc(e.note)}"`:''}><b>${dd(e.d)}.${dd(e.m)}</b>
     <span class="k k-${e.kind}">${KIND[e.kind]}</span>
     <span class="ti">${esc(e.name)}</span>
     <span class="in">${e.in===0?'сегодня':(e.in===1?'завтра':'через '+e.in+' дн.')}</span>
     ${e.amt||e.note?`<i class="pn">${e.amt?`<b>${esc(e.amt)}</b>`:''}${
       e.amt&&e.note?' · ':''}${esc(e.note)}</i>`:''}</div>`).join('')}
  </div>`:'<div class="hero none">В ближайший месяц событий нет.</div>'}
 <div class="cal">${order.map(m=>{
   const items=byMonth[m]||[];
   return `<div class="m${m===curM?' cur':''}${items.length?'':' mt'}">
    <h4>${MONT[m]}</h4>
    ${items.length?`<ul>${items.map(e=>`<li class="${e.in<=14?'soon':''} l-${e.kind}"${
      e.note?` title="${esc(e.note)}"`:''}>
      <b>${dd(e.d)}</b><span>${esc(e.name)}${e.amt?` - ${esc(e.amt)}`:''}</span></li>`).join('')}</ul>`
     :'<div class="none">-</div>'}</div>`;}).join('')}</div>`
 :'<div class="empty">Событий нет.</div>';
// Нажимаемой становится только та карточка, у которой комментарий действительно
// не поместился в одну строку. Это видно лишь после вёрстки - длина строки зависит
// от ширины экрана, - поэтому меряется здесь, а не решается в шаблоне. Иначе день
// рождения выглядел бы нажимаемым, а нажатие ничего бы не меняло.
document.querySelectorAll('#pane-bdays .hero-i .pn').forEach(n=>{
 if(n.scrollHeight>n.clientHeight+1)n.parentNode.classList.add('more');});
document.getElementById('pane-bdays').addEventListener('click',e=>{
 const n=e.target.closest('.hero-i.more'); if(n)n.classList.toggle('clip');});

// пояснения по вкладкам
const EXPL={
 bdays:[['Что в календаре','Даты платежей по графику сбора, дни рождения детей и учителя, даты мероприятий. Всё из журнала, вручную ничего не вносится.'],
  ['Ближайший месяц','Блок сверху показывает то, что произойдёт в течение ближайших недель, чтобы не листать календарь.'],
  ['Дни считаются от сегодня','«Через сколько дней» и подсветка ближайших двух недель берутся от настоящего сегодняшнего дня, а не от даты сборки: открыв ту же страницу через неделю, вы увидите обновившийся отсчёт. Деньги так не умеют - остаток, долги и доли посчитаны на дату отчёта, она указана под итогами.'],
  ['Планируемые события','Съёмки, экскурсии и прочее, что класс оплачивает мимо кассы: помечены «платят сами» и зелёным. Родители платят организатору напрямую, в кассу эти деньги не идут, на остаток и долю расходов не влияют, и долгом за такое событие никто не становится. Сумма в заголовке - цена с одного человека.'],
  ['Порядок месяцев','По учебному году, с сентября. Текущий месяц обведён рамкой, ближайшие две недели подсвечены жёлтым.']],
 sbory:[['Что такое сбор','Отдельная тема со своей суммой и своим списком участников. Кто в мероприятии не участвует, за него не платит и в его расходах не участвует.'],
  ['График платежей','Сумма за год разбита на части с датами. Планка «к сроку» - нарастающий итог по этому графику: все шаги, чья дата уже наступила на дату отчёта. Отдельно объявлять новую сумму не нужно, планка поднимается сама с каждой датой графика. В самом графике ближайший шаг выделен цветом, а пройденные показаны серым.'],
  ['Долг за год и долг к сроку','Главное число - долг за год: сколько осталось внести до полной суммы. Долг к сроку - его часть, которую ждут к ближайшей дате графика; она указана рядом. Закрыть ближайший срок не значит рассчитаться: за год может остаться ещё сумма, просто её срок пока не наступил.'],
  ['Оплата частями','Каждый платёж записывается отдельно, нехватка к сроку уменьшается на внесённую сумму.']],
 kids:[['Должен за год','Сколько ребёнку осталось внести до полной суммы сбора. Мелким шрифтом под суммой - какая её часть нужна к ближайшему сроку и какого числа. Если там «ближайший срок закрыт», ближайший платёж внесён, но за год сумма ещё остаётся.'],
  ['Не потрачено','Деньги ребёнка, которые ещё не ушли в расходы: внесено минус своя доля расходов. Расходы сбора делятся поровну между его участниками. Это не то же, что «долга нет»: можно ничего не быть должным и иметь остаток, а можно иметь остаток и долг одновременно.'],
  ['Минус в «не потрачено»','За ребёнка уже потрачено больше, чем он внёс. Это не ошибка: класс профинансировал его долю из общих денег.'],
  ['Что внутри строки','Нажмите на строку: там разбивка по деньгам, оставшиеся платежи по графику с датами, таблица по каждому сбору и список внесённых платежей.'],
  ['Считается по каждому сбору отдельно','Переплата по одному сбору не закрывает нехватку по другому. Нажмите на строку, чтобы увидеть разбивку по каждому сбору.'],
  ['Сортировка','Нажмите на заголовок колонки. Повторное нажатие меняет направление.'],
  ['Нашли ошибку','Если платежа нет или сумма не совпадает, напишите Ане, поправим и перевыпустим отчёт.']],
 spends:[['Обещано, но не оплачено','Услуга заказана или вещь обещана, а деньги ещё лежат на карте. В расходы такие строки не попадают и в долю на ребёнка не входят - доля считается только по тому, что уже потрачено. Когда оплата пройдёт, строка переедет в список трат. «Свободный остаток» в шапке - это остаток кассы за вычетом обещанного.'],
  ['Что именно купили','Если у расхода есть расшифровка - что вошло в чек, - строка раскрывается нажатием. Строки без расшифровки не раскрываются.'],
  ['Подтверждения','Бумажные чеки по классу не собираются. У каждого расхода указано, чем он подтверждён: скрин оплаты, чек или только со слов.'],
  ['Привязка к сбору','Каждый расход относится к конкретному сбору и делится только между его участниками.'],
  ['Направление','Категория расхода: праздник, подарки, канцтовары и так далее. По ней видно, на что уходят деньги класса.'],
  ['Строка с минусом','Возмещение: деньги вернулись в кассу - например, соседний класс возместил свою долю общей траты. Такая строка уменьшает и общие расходы, и долю каждого ребёнка поровну. Это не долг и не ошибка: денег в кассе стало больше, а не меньше.'],
  ['Группировка','Переключатель над списком: по датам, по направлениям или по сборам. Суммы в заголовках групп пересчитываются - в группе с возмещением итог может оказаться меньше отдельной траты или уйти в минус.']],
 flow:[['Зачем эта вкладка','Чтобы остаток не приходилось принимать на веру. Пройдите сверху вниз: каждая строка прибавляет или вычитает, справа - сколько осталось после неё. Последняя цифра - та же, что в плитке «Остаток» наверху.'],
  ['Порядок строк','Строго по дням, от первой операции к последней. Ничего не сортируется и не переставляется - иначе сквозной остаток потерял бы смысл.'],
  ['Месяцы','Свежий месяц раскрыт, прошедшие свёрнуты до строки итога: сколько пришло, сколько ушло и с чем месяц закончился. Нажмите на месяц, чтобы раскрыть его целиком.'],
  ['Взносы одной строкой','Платежи, пришедшие в один день, сведены в одну строку - «Взносы, N чел.». Кто именно и сколько внёс, видно во вкладке «По детям». Траты, наоборот, стоят каждая своей строкой.'],
  ['Возмещение в приходе','Возврат от соседнего класса за общее оформление стоит в приходе: в этот день на карте действительно стало больше. Во вкладке «Расходы» та же операция идёт со знаком минус - там знак нужен, чтобы возврат уменьшил и общие траты, и долю каждого ребёнка.']]};
// Пояснение появляется только когда личные списания есть: на пустом листе оно
// объясняло бы то, чего на странице нет.
if(PERS)EXPL.kids.splice(2,0,['Личные списания',
 'Расходную вещь, доставшуюся конкретному ребёнку, оплачивает он, а не весь класс: её стоимость входит в его долю расходов отдельной строкой. Доля тогда больше общей - у кого именно и за что, видно внутри строки. Это не долг: по графику взносов такой ребёнок должен ровно столько же, сколько все остальные, просто не потраченных денег у него на эту сумму меньше.']);
for(const [k,items] of Object.entries(EXPL)){
 const el=document.getElementById('ex-'+k); if(!el)continue;
 el.innerHTML=`<button class="btn wide" data-toggle="exp-${k}" aria-expanded="false">Как это считается<span class="chev">▾</span></button>
  <div class="panel" id="exp-${k}"><dl>${items.map(([a,b])=>`<dt>${a}</dt><dd>${b}</dd>`).join('')}</dl></div>`;}

// Обещанное, но не оплаченное: деньги ещё на карте, поэтому это не расходы и в
// доле на ребёнка их нет. Блок стоит выше списка трат и от группировки не зависит,
// поэтому собирается один раз. Комментарий раскрывается нажатием на строку.
const PROM=D.promises||[];
const promHtml=PROM.length?`<h2>Обещано, но ещё не оплачено · ${
  rub(PROM.reduce((s,p)=>s+p.amount,0))}</h2>
 <div class="rows">${PROM.map(p=>`<div class="row${p.note?'':' nc'}"><div class="row-h">
   ${p.date?`<span class="tag">${esc(p.date.slice(0,5))}</span>`:''}
   <span class="pm">${esc(p.what)}${p.who?`<span class="sub">${esc(p.who)}</span>`:''}</span>
   <span class="bal">${rub(p.amount)}</span>
   <span class="chev"${p.note?'':' style="visibility:hidden"'}>▾</span></div>
  ${p.note?`<div class="row-b"><div class="sum">${esc(p.note)}</div></div>`:''}</div>`).join('')}
 </div><h2>Уже потрачено</h2>`:'';

let SGROUP='date';
function renderSpends(){
 const el=document.getElementById('pane-spends');
 if(!D.spends.length){el.innerHTML=promHtml+'<div class="empty">Пока не потрачено ни рубля.</div>';return;}
 const key=e=>SGROUP==='date'?e.date:(SGROUP==='cat'?e.cat:e.sbor);
 const order=[...D.spends];
 // Дата сравнивается как ГГГГММДД, а не строкой «ДД.ММ.ГГГГ»: посимвольно первым
 // идёт день, и 01.09 встало бы раньше 31.08. Внутри групп по направлению и по
 // сбору ключ тот же - иначе на переходе через месяц порядок бы поехал.
 const dkey=s=>s.split('.').reverse().join('');
 if(SGROUP==='date')order.sort((a,b)=>dkey(b.date).localeCompare(dkey(a.date)));
 else order.sort((a,b)=>key(a).localeCompare(key(b))||dkey(a.date).localeCompare(dkey(b.date)));
 const gs=[];
 order.forEach(e=>{let g=gs[gs.length-1];
  if(!g||g.k!==key(e)){g={k:key(e),items:[],sum:0};gs.push(g);}
  g.items.push(e); g.sum+=e.amount;});
 const chips=[['date','По датам'],['cat','По направлениям'],['sbor','По сборам']].map(([k,l])=>
  `<button class="chip" data-sg="${k}" aria-pressed="${k===SGROUP}">${l}</button>`).join('');
 // Расход с комментарием раскрывается нажатием - тем же .row/.row-b, что и
 // обязательство. Без комментария тела нет, и строка не раскрывается.
 el.innerHTML=promHtml+`<div class="chips" id="sgroup">${chips}</div>
  ${gs.map(g=>`<div class="mgroup">
   <div class="mhead">${esc(g.k)}<span${g.sum<0?' class="back"':''}>${rub(g.sum)}</span></div>
   <div class="rows">${g.items.map(e=>`<div class="row${e.comment?'':' nc'}"><div class="row-h">
     ${SGROUP!=='date'?`<span class="tag">${esc(e.date)}</span>`:''}
     <span class="pm">${esc(e.what)}${(()=>{const t=[SGROUP!=='cat'?esc(e.cat):'',
       e.amount<0?'<b class="back">возмещение</b>':''].filter(Boolean);
       return t.length?`<span class="sub">${t.join(' · ')}</span>`:'';})()}</span>
     <span class="amt${e.amount<0?' back':''}">${rub(e.amount)}</span><span class="tag">${esc(e.proof)}</span>
     <span class="chev"${e.comment?'':' style="visibility:hidden"'}>▾</span></div>
    ${e.comment?`<div class="row-b"><div class="sum">${esc(e.comment)}</div></div>`:''}</div>`).join('')}
   </div></div>`).join('')}`;
 document.getElementById('sgroup').addEventListener('click',ev=>{
  const c=ev.target.closest('.chip'); if(!c)return; SGROUP=c.dataset.sg; renderSpends();});
}
renderSpends();

// Лента кассы. Смысл вкладки не в бухгалтерской форме, а в том, чтобы родитель
// пришёл к остатку из шапки сам, не принимая его на веру: «собрано» и «потрачено»
// лежат в разных списках, и свести их иначе нечем. Отсюда всё устройство: порядок
// строго хронологический, снизу вверх ничего не сортируется и не переставляется,
// а последняя строка обязана дать ту же цифру, что плитка «Остаток». Сходимость
// проверена на сборке (см. tools/build_report.py) и в tools/verify.py.
const FLOW=D.flow||[];
if(!FLOW.length){
 // Журнал без листа «Движение» - прошлогодний. Пятой вкладке тогда неоткуда
 // взяться, и пустой она висеть не должна: её просто нет.
 ['nav button[data-tab=flow]','#pane-flow','#ex-flow'].forEach(sel=>{
  const el=document.querySelector(sel); if(el)el.remove();});
}else{
 const fg=[];
 FLOW.forEach(f=>{const [d,m,y]=f.date.split('.').map(Number);
  let g=fg[fg.length-1];
  if(!g||g.m!==m||g.y!==y)fg.push(g={m:m,y:y,items:[],in:0,out:0,bal:0});
  g.items.push(f); g.in+=f.in; g.out+=f.out; g.bal=f.bal;});
 // Раскрыт последний месяц - тот, в котором лента заканчивается и стоит нужная
 // цифра. Прошлые свёрнуты до строки итога: к концу года операций наберётся за
 // сотню, и развёрнутая лента перестанет читаться.
 document.getElementById('pane-flow').innerHTML=fg.map((g,i)=>{
  const open=i===fg.length-1;
  return `<div class="mvm">
   <button class="btn mvh" data-toggle="fm${i}" aria-expanded="${open}">
    <span class="mn">${MONT[g.m-1]} ${g.y}</span>
    <span class="mb">осталось ${rub(g.bal)}</span><span class="chev">▾</span>
    <span class="ms">пришло ${rub(g.in)} · ушло ${rub(g.out)} · ${g.items.length} ${
      plural(g.items.length,'операция','операции','операций')}</span></button>
   <div class="panel${open?' open':''}" id="fm${i}"><div class="rows">
    <div class="mv mvhd"><span class="d">Дата</span><span class="op">Операция</span>
     <span class="am">Пришло / ушло</span><span class="bal">Осталось</span></div>${
    g.items.map(f=>`<div class="mv ${f.in?'plus':'minus'}">
     <span class="d">${esc(dm(f.date))}</span>
     <span class="op">${esc(f.what)}</span>
     <span class="am">${f.in?'+'+rub(f.in):rub(-f.out)}</span>
     <span class="bal"><i>осталось </i>${rub(f.bal)}</span></div>`).join('')}</div></div>
   </div>`;}).join('')
  +`<div class="hint">Это движение денег класса, а не выписка по карте казначея:
    часть покупок родители делали сами, а касса возмещала им переводом. Даты трат -
    это даты покупки, а не списания с карты. Строка «Взносы, N чел.» - платежи
    одного дня, сведённые в одну: по именам они разложены во вкладке «По детям».</div>`;
}

document.addEventListener('click',e=>{
 const tab=e.target.closest('nav button');
 if(tab){document.querySelectorAll('nav button').forEach(b=>b.setAttribute('aria-selected',String(b===tab)));
  // Крайняя вкладка на узком экране видна не целиком. Нажали по её краю - полоса
  // подтягивает её к себе, иначе выбранной оказывается кнопка, которую не видно.
  tab.scrollIntoView({block:'nearest',inline:'nearest'});
  // Вкладки перечислены здесь, а не собираются из nav: «Пришло и ушло» удаляется
  // из nav, когда листа «Движение» в журнале нет, и её панель уходит вместе с ней.
  ['bdays','sbory','kids','spends','flow'].forEach(n=>{
   const pn=document.getElementById('pane-'+n); if(pn)pn.hidden=(n!==tab.dataset.tab);
   const ex=document.getElementById('ex-'+n); if(ex)ex.hidden=(n!==tab.dataset.tab);});
  // «Как это считается» сворачивается при уходе с вкладки: вернувшись, человек
  // видит её в исходном виде, а не с развёрнутым посреди экрана пояснением
  document.querySelectorAll('.expl .panel.open').forEach(p=>{p.classList.remove('open');
   const b=document.querySelector(`[data-toggle="${p.id}"]`);
   if(b)b.setAttribute('aria-expanded','false');});
  const a=document.getElementById('navtop');
  if(window.scrollY>a.offsetTop-8)window.scrollTo({top:a.offsetTop-8});
  return;}
 const btn=e.target.closest('.btn[data-toggle]');
 if(btn){const p=document.getElementById(btn.dataset.toggle);
  // «Участники» и «Расходы» в одной карточке взаимоисключающи. Область поиска -
  // .card, поэтому соседний сбор не трогается, а кнопка «Как это считается»
  // лежит вне карточек и под это правило не попадает.
  const card=btn.closest('.card');
  if(card&&!p.classList.contains('open'))
   card.querySelectorAll('.panel.open').forEach(o=>{
    if(o===p)return;
    o.classList.remove('open');
    const ob=card.querySelector(`[data-toggle="${o.id}"]`);
    if(ob)ob.setAttribute('aria-expanded','false');});
  btn.setAttribute('aria-expanded',String(p.classList.toggle('open')));return;}
 const sh=e.target.closest('.sh');
 if(sh){const key=sh.dataset.sort;
  if(SORT===key)SDIR=-SDIR; else {SORT=key; SDIR=SDEF[key];}
  // перерисовка сбрасывает раскрытые панели и строки - запоминаем и возвращаем
  const open=[...document.querySelectorAll('#pane-sbory .panel.open')].map(p=>p.id);
  renderSbory(); open.forEach(id=>{const p=document.getElementById(id);
   if(p){p.classList.add('open');const b=document.querySelector(`[data-toggle="${id}"]`);
    if(b)b.setAttribute('aria-expanded','true');}});
  const rows=[...document.querySelectorAll('#kidlist .row.open')].map(r=>r.dataset.k);
  renderKids(document.getElementById('q').value);
  rows.forEach(k=>{const r=document.querySelector(`#kidlist .row[data-k="${k}"]`);
   if(r)r.classList.add('open');});
  return;}
 // Раскрывается только строка, у которой есть что показать: у обязательства без
 // комментария тела нет, и рамка «раскрыто» вокруг пустоты выглядела бы ошибкой.
 const rh=e.target.closest('.row-h');
 if(rh&&rh.parentElement.querySelector('.row-b'))rh.parentElement.classList.toggle('open');});

}

(function(){
 if(typeof RAW!=='string'||!RAW.startsWith('ENC:')){document.getElementById('gate').remove();
  document.getElementById('app').hidden=false; boot(RAW); return;}
 const g=document.getElementById('gate'), f=document.getElementById('gform'),
       inp=document.getElementById('gpass'), err=document.getElementById('gerr');
 f.addEventListener('submit', async ev=>{
  ev.preventDefault(); err.textContent=''; f.classList.add('busy');
  try{
   const D=await decryptPayload(RAW.slice(4), inp.value);
   g.remove(); const app=document.getElementById('app'); app.hidden=false; boot(D);
  }catch(e){
   f.classList.remove('busy'); err.textContent='Неверный код. Спросите у Ани.';
   inp.select();
  }});
 inp.focus();
})();
</script></body></html>
"""
PAGE = PAGE.replace("\u2014", "-").replace("\u2013", "-")
PAYLOAD = ("ENC:" + encrypt_payload(DATA, PASSWORD)) if PASSWORD else DATA
open(OUT, "w", encoding="utf-8").write(PAGE.replace("__DATA__", json.dumps(PAYLOAD, ensure_ascii=False) if PASSWORD else DATA)
                                        .replace("__ASOF__", AS_OF.strftime("%d.%m.%Y"))
                                        .replace("__TREAS__", TREASURER_LABEL)
                                        .replace("__SEARCHPH__", "Найти по фамилии или имени…")
                                        .replace("__STAMP__", STAMP)
                                        .replace("__ITERS__", str(PBKDF2_ITERS))
                                        .replace("__YEAR__", str(AS_OF.year))
                                        .replace("__MONTH__", str(AS_OF.month))
                                        .replace("__DAY__", str(AS_OF.day)))
# Учитель помечен как «не ученик», но плательщиком не является: в счётчике он
# идёт отдельно, иначе строка отчёта расходится с блоком поступлений на странице.
_ext = sum(1 for k in kids if k["ext"] and not k["teach"])
_tch = sum(1 for k in kids if k["teach"])
_pl = "внешний плательщик" if _ext % 10 == 1 and _ext % 100 != 11 else "внешних плательщиков"
_ts = "техническая строка" if _tch % 10 == 1 and _tch % 100 != 11 else "технических строк"
print("report ok ·", len(sbory), "сборов ·", len(kids) - _ext - _tch, "детей"
      + (f" · {_ext} {_pl}" if _ext else "")
      + (f" · {_tch} {_ts}" if _tch else ""))

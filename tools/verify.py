#!/usr/bin/env python3
"""Проверка собранной страницы: расшифровать payload, показать итоги, поймать утечку.

Запуск:

    PASSWORD='код' .venv/bin/python tools/verify.py index.html

Смысл в том, чтобы смотреть на то, что реально лежит в файле, а не на то, что
генератор напечатал в конце сборки. Расхождение между журналом и страницей
ловится глазами по итогам, а утечка персональных данных - проверками ниже.

Выход 0 - всё сошлось, 1 - хотя бы одна проверка упала.
"""

import base64
import json
import os
import re
import sys

from openpyxl import load_workbook

import journal

from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

HERE = os.path.dirname(os.path.abspath(__file__))
GEN = os.path.join(HERE, "build_report.py")

# Публичная форма имени - «Артур Е.»: имя, пробел, одна заглавная буква с точкой.
# Полная фамилия («Елисеев Артур») этот шаблон не проходит - на том и проверка.
MASKED = re.compile(r"^\S+ [А-ЯЁA-Z]\.$")

# Ключи payload, которые в файле обязаны быть только внутри шифротекста.
# Взяты те, что не встречаются в вёрстке и скриптах страницы: «"spends"» сюда
# не годится - так называется вкладка (data-tab="spends"), проверка ловила бы себя.
PLAIN_MARKERS = ('"debtY"', '"dueYear"', '"bday"', '"proof"', '"collected"', '"promises"')

# Служебные пометки казначея. В журнале они законны - это «здесь ещё не разобрались»,
# - но на страницу идти не должны: родитель читает их как часть отчёта.
MARKS = ("[?]", "уточнить", "уточняется", "TODO", "??")

ok = True


def check(cond, good, bad):
    """Печатает строку отчёта и запоминает провал, не прерывая остальные проверки."""
    global ok
    print(f"  {'ok  ' if cond else 'FAIL'}  {good if cond else bad}")
    if not cond:
        ok = False


def norm_pass(p):
    """Та же нормализация, что в build_report.py: иначе выведется другой ключ."""
    return p.strip().lower()


def gen_iters():
    """Число итераций из генератора - источник правды для страницы."""
    m = re.search(r"^PBKDF2_ITERS\s*=\s*(\d+)", open(GEN, encoding="utf-8").read(), re.M)
    if not m:
        sys.exit("не нашёл PBKDF2_ITERS в tools/build_report.py")
    return int(m.group(1))


def gen_const(pattern, what):
    """Константа из генератора. Читается оттуда, а не дублируется здесь: разъехавшись,
    копия молча увела бы проверку не на те строки журнала или не на ту подпись."""
    m = re.search(pattern, open(GEN, encoding="utf-8").read(), re.M)
    if not m:
        sys.exit(f"не нашёл {what} в tools/build_report.py")
    return m.groups()


def money(sheet, row, col):
    """Число из ячейки; пусто и текст считаются нулём - как num() в генераторе."""
    v = sheet.cell(row=row, column=col).value
    return v if isinstance(v, (int, float)) else 0


def roster():
    """Лист «Ученики» и границы списка - для проверок, которым нужен журнал.

    Границы берутся тем же journal.bounds(), что и в генераторе: разойдясь, копия
    молча увела бы проверку не на те строки. None означает, что журнал не передан -
    verify запускают и руками, над файлом, собранным раньше.
    """
    src = os.environ.get("SRC", "")
    if not src or not os.path.exists(src):
        return None
    uch = load_workbook(src, data_only=True)["Ученики"]
    first, last = journal.bounds(uch, "№")
    return uch, first, last


def count_check(data, spisok):
    """Сколько строк в списке класса - столько и на странице.

    Это проверка ровно того, чем оборачивается уехавшая граница списка: цикл по
    слишком короткому диапазону не падает, он молча не доходит до последних детей.
    Генератор такую потерю не заметит - у него в журнале ровно то, что он прочитал,
    - поэтому список пересчитывается здесь заново, от шапки до строки «ИТОГО».
    """
    uch, first, last = spisok
    rows = [r for r in range(first, last + 1) if str(uch.cell(row=r, column=3).value or "").strip()]
    ext = [r for r in rows
           if str(uch.cell(row=r, column=12).value or "").strip().lower().startswith("не ученик")]
    want, got = len(rows) - len(ext), len([k for k in data["kids"] if not k.get("ext")])
    check(want == got, f"детей на странице столько же, сколько в журнале ({got})",
          f"в журнале {want} детей (строки {first}-{last}), а на странице {got} - "
          f"часть списка потерялась по дороге")
    want_ext, got_ext = len(ext), len([k for k in data["kids"] if k.get("ext")])
    check(want_ext == got_ext,
          f"строк «Не ученик…» на странице столько же, сколько в журнале ({got_ext})",
          f"в журнале {want_ext} таких строк, а на странице {got_ext}")


def leak_check(html, data, spisok):
    """Технические строки листа «Ученики» - те, что помечены в колонке L как
    «Не ученик…», - бывают двух видов, и проверяются они по-разному.

    У кого есть деньги (соседний класс, скинувшийся на декор), имя стоит на
    странице по замыслу: без блока «Поступления не от учеников класса» сумма
    видимых остатков не сходится с остатком кассы. Такие строки под шаблон маски
    не подходят и из проверки формы имени исключены - как и было.

    У кого денег нет (учитель), имени на странице быть не должно вовсе: строка
    заведена ради дня рождения, а в колонке C может стоять настоящее ФИО.
    Проверяется не форма имени, а факт утечки - содержимое колонки C ищется и в
    расшифрованном payload, и в разметке. Это строже шаблона: шаблон поймал бы
    только неверно замаскированное имя, а поиск подстроки ловит любое попадание,
    включая тот день, когда кто-нибудь начнёт брать подпись из колонки C.

    Единственное исключение - когда колонка C дословно совпадает с подписью,
    которую печатает сам генератор: тогда совпадение даёт его собственный текст,
    а утекать в такой строке нечему."""
    uch, first, last = spisok
    label, = gen_const(r'^TEACHER_LABEL\s*=\s*"([^"]*)"', "TEACHER_LABEL")
    hay = (html + json.dumps(data, ensure_ascii=False)).casefold()
    checked, as_label, leaked = 0, 0, []
    for r in range(first, last + 1):
        name = str(uch.cell(row=r, column=3).value or "").strip()
        note = str(uch.cell(row=r, column=12).value or "").strip().lower()
        if not name or not note.startswith("не ученик"):
            continue
        # E - внесено, J - остаток. Деньги есть хотя бы в одной - это плательщик,
        # и его имя стоит на странице по замыслу.
        if money(uch, r, 5) or money(uch, r, 10):
            continue
        if name == label:
            as_label += 1                  # это и есть подпись из генератора
            continue
        checked += 1
        if name.casefold() in hay:
            leaked.append(f"строка {r}: «{name}»")
    if not checked:
        print(f'  ok    технические строки без денег: {as_label} шт., все названы '
              f'подписью «{label}» из кода - утекать нечему'
              if as_label else "  ok    технических строк без денег в журнале нет")
        return
    check(not leaked,
          f"колонка C технических строк на страницу не попала ({checked} шт.)",
          f"имя из журнала попало на страницу - {'; '.join(leaked)}")


def rub(n):
    """Число в том же виде, в каком его видит родитель: 9 810,5."""
    s = f"{round(n, 2):,.2f}".replace(",", " ").replace(".", ",")
    return s[:-3] if s.endswith(",00") else s


def find_marks(node, path=""):
    """Обходит payload и возвращает [(путь, пометки, текст)] для каждого поля
    со служебной пометкой. Путь ведёт до конкретной записи и поля - иначе
    в журнале на пару сотен строк непонятно, где именно править."""
    if isinstance(node, str):
        hit = [m for m in MARKS if m in node]
        return [(path, hit, node)] if hit else []
    if isinstance(node, list):
        return [h for i, v in enumerate(node) for h in find_marks(v, f"{path}[{i}]")]
    if isinstance(node, dict):
        # Запись подписываем её собственным названием: «spends[3] «Хозтовары в класс»»
        # читается, а «spends[3]» пришлось бы искать глазами.
        label = next((str(node[k]) for k in ("what", "title", "sbor", "name") if node.get(k)), "")
        here = f"{path} «{label}»" if label and path else path
        return [h for k, v in node.items() for h in find_marks(v, f"{here}.{k}" if here else k)]
    return []


def decrypt(blob_b64, password, iters):
    blob = base64.b64decode(blob_b64)
    salt, iv, ct = blob[:16], blob[16:28], blob[28:]
    key = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt,
                     iterations=iters).derive(norm_pass(password).encode())
    return json.loads(AESGCM(key).decrypt(iv, ct, None))


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "index.html"
    password = os.environ.get("PASSWORD", "")
    if not password:
        sys.exit("нужен PASSWORD - без кода payload не расшифровать")

    html = open(path, encoding="utf-8").read()
    iters = gen_iters()

    print(f"\n{path}\n")
    print("Шифрование")
    m = re.search(r'const RAW="ENC:([A-Za-z0-9+/=]+)"', html)
    check(m is not None, "данные зашифрованы (RAW=\"ENC:...\")",
          "данных в зашифрованном виде нет - страницу нельзя публиковать")
    if not m:
        sys.exit(1)

    page_iters = [int(x) for x in re.findall(r"iterations:(\d+)", html)]
    check(page_iters and all(i == iters for i in page_iters),
          f"итерации PBKDF2 совпадают с генератором ({iters})",
          f"итерации разошлись: в генераторе {iters}, на странице {page_iters}")

    data = decrypt(m.group(1), password, iters)
    kids, sbory, spends, t = data["kids"], data["sbory"], data["spends"], data["t"]
    # Обязательства лежат отдельным ключом и в spends не подмешиваются: иначе
    # «потрачено» и доля на ребёнка выросли бы на ещё не ушедшие с карты деньги.
    promises = data.get("promises", [])
    # Внешние плательщики (ext) лежат в том же списке, но детьми класса не являются:
    # расходы делятся не на них, и в счёт детей они не идут.
    own = [k for k in kids if not k.get("ext")]
    ext = [k for k in kids if k.get("ext")]
    # Учитель тоже «не ученик», но плательщиком не является: денег за ним нет,
    # и в блок поступлений он не идёт. В счётчике их смешивать нельзя.
    payers = [k for k in ext if not k.get("teach")]
    tech = [k for k in ext if k.get("teach")]

    print("\nСодержимое")
    print(f"        детей: {len(own)}")
    if payers:
        print(f"        внешних плательщиков: {len(payers)}")
    if tech:
        print(f"        технических строк: {len(tech)}")
    print(f"        сборов: {len(sbory)}")
    print(f"        расходов: {len(spends)}")
    if promises:
        print(f"        обязательств: {len(promises)}")
    print(f"        собрано: {rub(t['collected'])} ₽")
    print(f"        потрачено: {rub(t['spent'])} ₽")
    print(f"        остаток: {rub(t['rest'])} ₽")
    if t.get("promised"):
        print(f"        обещано: {rub(t['promised'])} ₽")
        print(f"        свободный остаток: {rub(t.get('free', 0))} ₽")
    if own:
        print(f"        доля расходов: {rub(t['spent'] / len(own))} ₽ на ребёнка")

    print("\nАрифметика")
    check(abs(t["collected"] - t["spent"] - t["rest"]) < 0.005,
          "собрано − потрачено = остаток",
          f"{rub(t['collected'])} − {rub(t['spent'])} ≠ {rub(t['rest'])}")
    total = sum(s["amount"] for s in spends)
    check(abs(total - t["spent"]) < 0.005,
          f"сумма {len(spends)} расходов сходится с «потрачено»",
          f"расходы дают {rub(total)}, а в итогах {rub(t['spent'])}")
    promised, free = t.get("promised", 0), t.get("free", t["rest"])
    if promises or promised:
        total_prom = sum(p["amount"] for p in promises)
        check(abs(total_prom - promised) < 0.005,
              f"сумма {len(promises)} обязательств сходится с «обещано»",
              f"обязательства дают {rub(total_prom)}, а в итогах {rub(promised)}")
        check(abs(t["rest"] - promised - free) < 0.005,
              "остаток − обещанное = свободный остаток",
              f"{rub(t['rest'])} − {rub(promised)} ≠ {rub(free)}")
        # Обязательство, случайно продублированное в «Расходах», ловится здесь:
        # оно ушло бы и в «потрачено», и в долю на ребёнка.
        dup = sorted({p["what"] for p in promises} & {s["what"] for s in spends})
        check(not dup, "обязательства не задваиваются с расходами",
              f"есть и в обязательствах, и в расходах: {', '.join(dup)}")
    if own:
        share = t["spent"] / len(own)
        bad = [k for k in own if abs(k.get("share", 0) - share) > 0.005]
        check(not bad, f"доля расходов у всех детей одна ({rub(share)} ₽)",
              f"доля расходов разъехалась у {len(bad)} детей")

    # Проверки, сверяющие страницу с журналом, требуют SRC: verify запускают и
    # руками, над файлом, собранным раньше. Список читается один раз на обе.
    spisok = roster()
    no_src = "        сборка через ./tools/build.sh журнал передаёт - см. README"

    print("\nСписок класса")
    if spisok is None:
        print("  ....  список не сверялся с журналом: нет журнала (SRC)")
        print(no_src)
    else:
        count_check(data, spisok)

    print("\nПерсональные данные")
    # Имена внешних плательщиков - не фамилии детей («Соседи»), под шаблон маски
    # они не подходят и проверяются не здесь. Собранное без MASK=1 всё равно
    # ловится: у детей фамилии остались бы полными.
    ext_names = {k["name"] for k in ext}
    names = [k["name"] for k in own]
    names += [p["name"] for s in sbory for p in s.get("parts", []) if p["name"] not in ext_names]
    leaked = sorted({n for n in names if not MASKED.match(n)})
    check(not leaked, f"все имена в payload замаскированы ({len(names)} шт.)",
          f"незамаскированных имён: {len(leaked)} - собрано без MASK=1")

    if spisok is None:
        print("  ....  колонка C технических строк не проверена: нет журнала (SRC)")
        print(no_src)
    else:
        leak_check(html, data, spisok)

    treas = re.search(r"казначей ([^<]+)</div>", html)
    check(treas is not None and MASKED.match(treas.group(1).strip()) is not None,
          "имя казначея замаскировано",
          f"имя казначея в открытом виде: {treas.group(1).strip() if treas else '—'}")

    found = [mk for mk in PLAIN_MARKERS if mk in html]
    check(not found, "открытых данных payload в файле нет",
          f"payload лежит в открытом виде: {', '.join(found)}")

    print("\nСлужебные пометки")
    marks = find_marks(data)
    check(not marks, f"на странице нет пометок «{'», «'.join(MARKS)}»",
          f"пометки увидят родители - {len(marks)} шт., чинить в журнале:")
    for path, hit, text in marks:
        print(f"          {path} — {', '.join(hit)}")
        print(f"            {text.strip()[:150]}")

    print("\n" + ("verify ok" if ok else "verify FAILED") + "\n")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
